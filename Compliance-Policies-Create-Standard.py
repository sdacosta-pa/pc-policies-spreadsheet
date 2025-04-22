import json
import csv
import requests
import aiohttp
import asyncio
import configparser
import shutil
import logging
from io import StringIO
from collections import defaultdict

import re
# import threading
import time
import datetime
from datetime import datetime, timedelta
from colorama import Fore, Style, init

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

# Specify your API config file
API_CONFIG_PATH = 'API_config.ini'

SSL_VERIFY = False
# include_AccountGroups = False
include_AccountGroups = True  # this will generate a lot of data for csv for if large number of rules and/or account groups are present

class ColorFormatter(logging.Formatter):
    """Custom formatter to add colors to log levels."""
    COLORS = {
        logging.DEBUG: Fore.CYAN,
        logging.INFO: Fore.GREEN,
        logging.WARNING: Fore.YELLOW,
        logging.ERROR: Fore.RED,
        logging.CRITICAL: Fore.MAGENTA,
    }

    def format(self, record):
        """Format log messages with colors and timestamps."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Manual timestamp
        log_color = self.COLORS.get(record.levelno, Fore.WHITE)
        message = super().format(record)  # Get the formatted log message
        return f"{log_color}{timestamp} - {record.levelname} - {record.funcName} - {record.lineno} - {message}{Style.RESET_ALL}"

# Initialize colorama
init(autoreset=True)

# Create a handler with the custom formatter
handler = logging.StreamHandler()
formatter = ColorFormatter("%(message)s")  # Keep format string minimal
handler.setFormatter(formatter)

# Configure logging correctly
logger = logging.getLogger()  # Get the root logger
logger.setLevel(logging.DEBUG)
logger.addHandler(handler)  # Add the custom handler
# logging.basicConfig(
#     # level=logging.INFO,
#     level=logging.DEBUG,
#     handlers=[handler],
# )
# Adjust the log level for httpx and httpcore to suppress their DEBUG logs.
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)


start_time = time.time()


def read_api_config():
    config = configparser.ConfigParser()
    config.read(API_CONFIG_PATH)
    baseurl = config.get('URL', 'BaseURL')
    access_key_id = config.get('AUTHENTICATION', 'ACCESS_KEY_ID')
    secret_key = config.get('AUTHENTICATION', 'SECRET_KEY')
    # print(f"{baseurl} {access_key_id} {secret_key}")
    return baseurl, access_key_id, secret_key


class TokenManager:
    def __init__(self, baseurl, access_key_id, secret_key, session):  # Add session parameter
        self.baseurl = baseurl
        self.access_key_id = access_key_id
        self.secret_key = secret_key
        self.token = None
        self.token_expiry_time = 0
        self.lock = asyncio.Lock()

    async def get_token(self):
        async with self.lock:
            if not self.token or time.time() >= self.token_expiry_time:
                await self._refresh_token()
            return self.token

    async def _refresh_token(self):
        url = f"{self.baseurl}/login"
        headers = {'Content-Type': 'application/json'}
        data = {"username": self.access_key_id, "password": self.secret_key}
        async with aiohttp.ClientSession() as session:
            async with session.post(url, headers=headers, json=data, ssl=SSL_VERIFY) as response:
                if response.status == 200:
                    response_data = await response.json()
                    self.token = response_data.get('token')
                    self.token_expiry_time = time.time() + 480  # Token valid for 8 minutes
                else:
                    raise Exception(f"Failed to get access token: {await response.text()}")


def sanitize_value(value):
    """Sanitize a value to replace illegal characters for Excel, including adding line breaks."""
    if isinstance(value, str):
        # Replace newline characters (\n) with the Excel line break (\r\n)
        sanitized_value = value.replace('\n', '\r\n')
        # Remove other control characters
        sanitized_value = re.sub(r'[\x00-\x09\x0B\x0C\x0E-\x1F\x7F]', ' ', sanitized_value)
        return sanitized_value.strip()
    return value


async def make_get_request(url, token_manager, session, max_retries=3, backoff_factor=5):
    """Send a GET request with retry logic and return the JSON response asynchronously."""
    token = await token_manager.get_token()
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    for attempt in range(1, max_retries + 1):
        try:
            async with session.get(url, headers=headers, ssl=SSL_VERIFY) as response:
                response_text = await response.text()
                if response.status == 200:
                    return await response.json()
                else:
                    print(Fore.RED + f"Failed to fetch data {response.status}: {url} {response_text}")
                    print(Fore.WHITE)
                    if attempt == max_retries:
                        return response.status
                    await asyncio.sleep(backoff_factor ** (attempt - 1))
        except aiohttp.ClientError as e:
            print(Fore.RED + f"Error during API call: {e}. URL: {url}, Attempt {attempt} of {max_retries}")
            print(Fore.WHITE)
            if attempt == max_retries:
                return None
            await asyncio.sleep(backoff_factor ** (attempt - 1))
        except asyncio.TimeoutError:
            print(Fore.RED + f"Request timed out. URL: {url}, Attempt {attempt} of {max_retries}")
            print(Fore.WHITE)
            if attempt == max_retries:
                return None
            await asyncio.sleep(backoff_factor ** (attempt - 1))
        except Exception as e:
            print(Fore.RED + f"Unexpected error: {e}. URL: {url}, Attempt {attempt} of {max_retries}")
            print(Fore.WHITE)
            if attempt == max_retries:
                return None
            await asyncio.sleep(backoff_factor ** (attempt - 1))


async def make_post_request(url, token_manager, payload, session, max_retries=3, backoff_factor=5):
    token = await token_manager.get_token()
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    for attempt in range(1, max_retries + 1):
        try:
            async with session.post(url, headers=headers, json=payload, ssl=SSL_VERIFY) as response:
                if response.status == 200:
                    return response.status
                else:
                    print(Fore.RED + f"Failed to create data {response.status}: {url}")
                    print(Fore.WHITE)
                    if attempt == max_retries:
                        return response.status
                    await asyncio.sleep(backoff_factor ** (attempt - 1))
        except Exception as e:
            print(Fore.RED + f"Unexpected error: {e}. URL: {url}, Attempt {attempt} of {max_retries}")
            print(Fore.WHITE)
            if attempt == max_retries:
                return None
            await asyncio.sleep(backoff_factor ** (attempt - 1))


async def make_put_request(url, token_manager, payload, session, max_retries=3, backoff_factor=5):
    token = await token_manager.get_token()
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    for attempt in range(1, max_retries + 1):
        try:
            async with session.put(url, headers=headers, json=payload, ssl=SSL_VERIFY) as response:
                if response.status == 200:
                    return response.status
                else:
                    print(Fore.RED + f"Failed to update data {response.status}: {url}")
                    print(Fore.WHITE)
                    if attempt == max_retries:
                        return response.status
                    await asyncio.sleep(backoff_factor ** (attempt - 1))
        except Exception as e:
            print(Fore.RED + f"Unexpected error: {e}. URL: {url}, Attempt {attempt} of {max_retries}")
            print(Fore.WHITE)
            if attempt == max_retries:
                return None
            await asyncio.sleep(backoff_factor ** (attempt - 1))

async def make_delete_request(url, token_manager, session, max_retries=3, backoff_factor=5):
    token = await token_manager.get_token()
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    for attempt in range(1, max_retries + 1):
        try:
            async with session.delete(url, headers=headers, ssl=SSL_VERIFY) as response:
                if response.status == 200:
                    return response.status
                else:
                    print(f"Failed to delete data {response.status}: {url}")
                    if attempt == max_retries:
                        return response.status
                    await asyncio.sleep(backoff_factor ** (attempt - 1))
        except Exception as e:
            print(f"Unexpected error: {e}. URL: {url}, Attempt {attempt} of {max_retries}")
            if attempt == max_retries:
                return None
            await asyncio.sleep(backoff_factor ** (attempt - 1))

# Function to write CSV to a local file
def write_csv_to_file(file_key, data):
    file_path = f"{file_key}"
    with open(file_path, 'w', newline='') as file:
        writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
        writer.writerows(data)
    print(f"CSV file saved to local path {file_path}")


# Function to read Excel data from a local file
def read_excel_from_file(folder, file_key):
    file_path = f"{folder}/{file_key}"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data


# Function to write Excel to a local file
def write_excel_to_file(folder, file_key, data):
    file_path = f"{folder}/{file_key}"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row_num, row in enumerate(data, start=1):
        for col_num, cell_value in enumerate(row, start=1):
            sanitized_value = sanitize_value(cell_value)  # Assuming you have a sanitize_value function
            sheet.cell(row=row_num, column=col_num, value=sanitized_value)
    workbook.save(file_path)
    print(f"Excel file saved to local path {file_path}")


async def prompt_for_standard_selection(standards_data):
    """Prompt the user to select a compliance standard, sorted by policiesAssignedCount, and return the selected standard details."""
    # Sort standards by policiesAssignedCount in descending order
    sorted_standards = sorted(standards_data, key=lambda x: x.get("name", 0), reverse=False)
    print("Available Compliance Standards (sorted by policies assigned):")
    for index, standard in enumerate(sorted_standards, start=1):
        print(Fore.GREEN + f"{index}. {standard.get('name')}" + Fore.WHITE + f" - Policies Assigned: {standard.get('policiesAssignedCount')}, "
              f"cloudType: {', '.join(standard.get('cloudType', []))}, createdBy: {standard.get('createdBy')}")
    # Prompt the user for a selection
    try:
        selection = int(input("Select a compliance standard by number: ")) - 1
        if selection < 0 or selection >= len(sorted_standards):
            print("Invalid selection. Please try again.")
            return await prompt_for_standard_selection(standards_data)  # Retry if invalid
        selected_standard = sorted_standards[selection]
        # Extract the cloud types from the selected standard
        stnd_id = selected_standard.get("id")
        cloud_types = selected_standard.get("cloudType", [])
        print(f"Selected Standard ID: {stnd_id}")
        print(f"Cloud Types for this standard: {cloud_types}")
        return selected_standard
    except ValueError:
        print("Invalid input. Please enter a number.")
        return await prompt_for_standard_selection(standards_data)


async def generate_compliance_summary(baseurl, token_manager, standard_id, session):
    # Dictionary to store the summary data
    summary_data = [["Section ID", "Section Description", "Policies Assigned Count", "Requirement ID", "Requirement Name", "Requirement Description"]]
    requirement_url = f"{baseurl}/compliance/{standard_id}/requirement"
    requirements_data = await make_get_request(requirement_url, token_manager, session)

    if not requirements_data:
        print("No requirements found for the standard.")
        return summary_data

    for requirement in requirements_data:
        requirement_id = requirement.get("id")
        requirement_Id = requirement.get("requirementId")
        requirement_name = requirement.get("name")
        requirement_description = requirement.get("description")

        section_url = f"{baseurl}/compliance/{requirement_id}/section"
        section_data = await make_get_request(section_url, token_manager, session)

        if not section_data:
            print(f"No sections found for requirement ID: {requirement_id}")
            continue

        for section in section_data:
            section_id = section.get("sectionId")
            section_description = section.get("description")
            policies_assigned_count = section.get("policiesAssignedCount", 0)

            summary_data.append([
                section_id,
                section_description,
                policies_assigned_count,
                requirement_Id,
                requirement_name,
                requirement_description
            ])
    return summary_data


def generate_csv(customer_name, stnd_name, totals, maybe_totals, included_totals, custom_totals, alert_rule_totals, auto_dismiss_totals, alert_rules_dict):
    output = StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    # Write totals for policies and custom policies
    writer.writerow(["CloudType", "Included", "Policies", "Custom"])
    total_counts = defaultdict(int)
    for cloud_type in sorted(totals.keys()):
        included_data = totals[cloud_type]
        for included in sorted(included_data.keys()):
            count = included_data[included]
            custom_count = custom_totals[cloud_type].get(included, 0)
            # Replace `included` value with `stnd_name` if it matches `customer_name` for display purposes
            display_included = stnd_name if included == customer_name else included
            writer.writerow([cloud_type, display_included, count, custom_count])
            total_counts[included] += count
            total_counts[f'{included}_Custom'] += custom_count
    writer.writerow([])
    writer.writerow([f"{customer_name} Total Policies for all clouds"])
    for included, count in total_counts.items():
        if included.endswith("_Custom"):
            continue
        # Replace `included` value with `stnd_name` if it matches `customer_name` for display purposes
        display_included = stnd_name if included == customer_name else included
        writer.writerow(["All Clouds", f"Total {display_included}", count])
    total_policies = sum(total_counts[included] for included in total_counts if not included.endswith("_Custom"))
    writer.writerow(["All Clouds", "Total Policies", total_policies])
    print(f"{customer_name} Total Policies: {total_policies}")
    writer.writerow(["All Clouds", "Total Custom", sum(total_counts[included] for included in total_counts if included.endswith("_Custom"))])
    # Write totals for MAYBE and YES policies per cloud
    # writer.writerow([])
    # writer.writerow([f"Compliance Standard - {stnd_name}"])
    # writer.writerow(["CloudType", "Requirement/Section", "Value", "MAYBE Policies", "Included Policies"])
    for cloud_type in sorted(maybe_totals.keys() | included_totals.keys()):  # Ensure all cloud types are considered
        maybe_data = maybe_totals.get(cloud_type, {})
        included_data = included_totals.get(cloud_type, {})
        all_categories = set(maybe_data.keys()) | set(included_data.keys())  # Combine keys from both dictionaries
        for category in sorted(all_categories):  # Process each category
            maybe_values = maybe_data.get(category, {})
            included_values = included_data.get(category, {})
            all_values = set(maybe_values.keys()) | set(included_values.keys())  # Combine all values in the category
            for value in sorted(all_values):  # Process each value
                maybe_count = maybe_values.get(value, 0)  # Get count from maybe_totals
                included_count = included_values.get(value, 0)  # Get count from included_totals
                writer.writerow([cloud_type, category, value, maybe_count, included_count])
    writer.writerow([])

    # Write alert rule totals
    # Write detailed information for auto dismiss rules
    auto_dismiss_rules, total_unique_resource_list_auto_dismiss, total_auto_dismiss_rules = get_unique_resource_lists_count_for_auto_dismiss(alert_rules_dict)
    writer.writerow([])
    writer.writerow(["Alert Rules"])
    writer.writerow(["Alert Rule Name", "Policies"])
    for rule, count in alert_rule_totals.items():
        writer.writerow([rule, count])
    writer.writerow(["Auto Dismiss Rules:"])
    for rule in auto_dismiss_rules:
        writer.writerow([rule])
    writer.writerow(["Total Active Auto Dismiss Alert Rules", auto_dismiss_totals])
    writer.writerow(["Total Tag-based Resource Lists in Auto Dismiss Rules", total_unique_resource_list_auto_dismiss])
    return output.getvalue()



async def get_requirements(baseurl, token_manager, standard_id, p_sections, session):
    # Dictionary to store mapping of p_section to (requirementName, compliance_id)
    requirements_dict = {}
    # Fetch all requirements for the given standard_id
    requirement_url = f"{baseurl}/compliance/{standard_id}/requirement"
    requirements_data = await make_get_request(requirement_url, token_manager, session)
    # Extract requirement IDs from the data
    requirement_ids = [req.get('id') for req in requirements_data if req.get('id')]
    # print(f"Requirement IDs: {requirement_ids}")
    # Convert p_sections to strings to ensure consistent comparison
    p_sections = [str(p).strip() for p in p_sections]
    # Iterate over each requirement_id
    for requirement_id in requirement_ids:
        # Fetch sections for the current requirement
        section_url = f"{baseurl}/compliance/{requirement_id}/section"
        section_data = await make_get_request(section_url, token_manager, session)
        # print(f"Section data for requirement ID {requirement_id}: {section_data}")
        # Process section_data as a list of dictionaries
        for section in section_data:
            section_id = str(section.get('sectionId', '')).strip()  # Convert sectionId to string
            requirement_name = section.get('requirementName', '').strip()
            compliance_id = section.get('id')
            # Debug each potential match
            # print(f"Checking if sectionId '{section_id}' matches any p_section: {p_sections}")
            # Match sectionId with p_sections
            if section_id in p_sections:  # Both are strings, comparison will work
                # print(f"Matched sectionId: {section_id} with p_sections: {p_sections}")
                requirements_dict[section_id] = (requirement_name, compliance_id)
        # Debug output for the final dictionary
        # for p_section, (requirement_name, compliance_id) in requirements_dict.items():
            # print(f"Matched Section ID: {p_section}, Requirement: {requirement_name}, Compliance ID: {compliance_id}")
    return requirements_dict

async def update_requirements(baseurl, token_manager, row_data, sheet_headers, customer_name, standard_id, stnd_name, session):
    # Dictionary to store mapping of sp_code to (requirement_id, requirement_name, section_description)
    standard_dict = {}

    # Keep track of updated requirements
    updated_requirements = set()

    for row in row_data:  # Iterate through each row in row_data
        # Extracting required values from the row
        requirement_name = row[sheet_headers.index('Requirement')]
        requirement_id = row[sheet_headers.index('Requirement_ID')]
        section_name = row[sheet_headers.index('section')]
        section_description = f"{requirement_id} - {section_name}"

        # Fetch all requirements for the given standard_id
        requirement_url = f"{baseurl}/compliance/{standard_id}/requirement"
        requirements_data = await make_get_request(requirement_url, token_manager, session)

        # Extract requirement IDs and names from the data
        requirements = [
            {"id": req.get("id"), "name": req.get("name")}
            for req in requirements_data
            if req.get("id") and req.get("name")
        ]
        requirement_ids = {req["id"] for req in requirements_data if "id" in req}

        if requirement_id not in requirement_ids:
            # Create requirements from CSV and create sections
            print(f"Creating requirement: {requirement_id}")

            updated_requirements_status = await create_requirements(
                baseurl, token_manager, requirement_name, requirement_id, customer_name,
                standard_id, stnd_name, session
            )
            if updated_requirements_status == '200':
                requirements.append({"id": requirement_id, "name": requirement_name})
            else:
                logging.error(f"Failed to create requirement {requirement_id}: {updated_requirements_status}")
        else:
            logging.info(f"Requirement {requirement_id} already exists, skipping creation.")
        # Iterate over each requirement
        for requirement in requirements:
            requirement_id = requirement["id"]
            requirement_name = requirement["name"]
            if requirement_name == requirement_name:
            # Update the requirement only if it hasn't been updated yet
                if (requirement_name, requirement_id) not in updated_requirements:
                    print(f"Updating requirement: {requirement_id}")

                    # URL for updating the requirement
                    requirements_url = f"{baseurl}/compliance/requirement/{requirement_id}"
                    requirement_payload = {
                        "description": requirement_name,
                        "name": requirement_name,
                        "requirementId": requirement_id
                    }

                    # Update the requirement
                    req_updated_status = await make_put_request(requirements_url, token_manager, requirement_payload, session)
                    if str(req_updated_status) == '200':
                        print(f"Requirement updated successfully: {requirement_id} {requirement_payload}")
                        updated_requirements.add((requirement_name, requirement_id))
                    else:
                        print(f"Failed to update requirement: {requirement_id}. Status: {req_updated_status} {requirement_payload}")
            # **Check if the section exists before creating it**
            section_url = f"{baseurl}/compliance/{requirement_id}/section"
            existing_sections = await make_get_request(section_url, token_manager, session)
            existing_section_ids = {section["sectionId"] for section in existing_sections if "sectionId" in section}

            if section_name not in existing_section_ids:
                # Create the section for each row
                section_url = f"{baseurl}/compliance/{requirement_id}/section"
                section_payload = {
                    "description": section_description,
                    "sectionId": section_name
                }

                print(f"Creating section: {requirement_id} with section_name: {section_name}")
                sec_updated_status = await make_post_request(section_url, token_manager, section_payload, session)

                if str(sec_updated_status) == '200':
                    print(f"Section created successfully for {requirement_id} with section_name: {section_name} {section_payload}")
                    standard_dict[requirement_id] = (requirement_name, section_name, section_description)
                else:
                    print(f"Failed to create section for {requirement_id} with sp_code: {section_name}. Status: {sec_updated_status} {section_payload}")
            else:
                logging.info(f"Section already exists for {requirement_id} with section_name: {section_name}, skipping creation.")

    return standard_dict

async def create_requirements(baseurl, token_manager, requirement_name, requirement_id, customer_name, standard_id, stnd_name, session):

    # URL for updating the requirement
    requirements_add_url = f"{baseurl}/compliance/{standard_id}/requirement"
    requirement_payload = {
        "requirementId": requirement_id,
        "name": requirement_name
    }
    # Update the requirement
    req_updated_status = await make_post_request(requirements_add_url, token_manager, requirement_payload, session)
    if str(req_updated_status) == '200':
        print(f"Requirement created successfully: {requirement_id} {requirement_payload}")
    else:
        print(f"Failed to create requirement: {requirement_id}. Status: {req_updated_status} {requirement_payload}")

    return req_updated_status


async def main():
    timestamp = datetime.now().strftime('%Y%m%d%H%M')

    test_event = {
        "Customer": "Widget",
        "bucket_name": "data",
        "filename": "compliance_standard",


        # "secret_name": "compliance-policy"
    }
    baseurl, access_key_id, secret_key = read_api_config()
    async with aiohttp.ClientSession() as session:

        token_manager = TokenManager(baseurl, access_key_id, secret_key, session)
        customer_name = test_event.get('Customer', 'Customer_Test')
        customer_ID = customer_name[:2].upper()  # Get the first two letters of Customer name in capitals
        standard_id = test_event.get('standard_id', '')
        customer_label = customer_name.replace(' ', '_')
        standard_id = test_event.get('standard_id', '')

        bucket_name = test_event.get('bucket_name', 'pc-compliance-pl')  # this is folder name
        filename = test_event.get('filename', 'compliance_policies')

        if standard_id:
            stnd_url = f"{baseurl}/compliance/{standard_id}"
            stnd_info = await make_get_request(stnd_url, token_manager, session)
            if not stnd_info:
                return []
            stnd_name = stnd_info.get('name', 'Unknown Standard')
            cloud_types = stnd_info.get("cloudType", [])
        else:
            standards_url = f"{baseurl}/compliance"
            standards_data = await make_get_request(standards_url, token_manager, session)
            # Filter the data for entries where "systemDefault" is False
            filtered_standards_data = [standard for standard in standards_data if not standard.get("systemDefault", True)]
            # Prompt the user to select a compliance standard
            selected_standard = await prompt_for_standard_selection(filtered_standards_data)
            if not selected_standard:
                print("No compliance standard selected.")
                return
            # Extract the stnd_id and cloud_types from the selected standard
            standard_id = selected_standard.get("id")
            # cloud_types = selected_standard.get("cloudType", [])
            # cloud_types = ['azure', 'aws'] # Override manually if required
            cloud_types = ['aws'] # Override manually if required
            stnd_name = selected_standard.get('name')
        stnd_label = stnd_name.replace(' ', '_')
        print(f"Selected Standard {stnd_name} - ID: {standard_id}")
        print(f"Cloud Types for this standard: {cloud_types}")
        print(f"Processing {customer_name} standard for {stnd_name}")
        # Get compliance metadata for the given standard_id
        compliance_url = f"{baseurl}/policy/compliance"
        # compliance_metadata = await make_get_request(compliance_url, token_manager, session)
        input_file_key = f"{bucket_name}/{filename}_{customer_label}_{stnd_label}_latest.xlsx"
        consolidated_file_key = f"{bucket_name}/{filename}_{customer_label}_{stnd_label}_data_totals_{timestamp}.csv"
        # updated_file_key = f"{bucket_name}/{filename}_{customer_label}_{stnd_label}_latest_{timestamp}.xlsx"
        standard_file_key = f"{bucket_name}/{filename}_{customer_label}_{stnd_label}_Reference_latest.csv"


        # Load workbook and get the sheet
        workbook = openpyxl.load_workbook(input_file_key)
        sheet = workbook.active
        import_data = []
        for row in sheet.iter_rows(values_only=True):
            import_data.append(list(row))
        sheet_headers = import_data[0]
        rows = import_data[1:]
        print(f"Standard {sheet_headers}")
        # print(f"Standard {rows}")
        # Error cleanup
        # await clean_adf_sections(baseurl, token_manager, standard_id, session)
        #
        # # # Update/create requirements and create sections
        standard_dict = await update_requirements(
            baseurl, token_manager, rows, sheet_headers, customer_name,
            standard_id, stnd_name, session
        )
        print(f"Standard {customer_name}: {standard_dict}")
        standard_summary_data = await generate_compliance_summary(baseurl, token_manager, standard_id, session)
        # print(standard_summary_data)
        if standard_summary_data:
            write_csv_to_file(standard_file_key, standard_summary_data)
    # Record the end time
    end_time = time.time()
    # Calculate the elapsed time
    elapsed_time_seconds = end_time - start_time
    # Convert elapsed time to minutes
    elapsed_time_minutes = elapsed_time_seconds / 60
    # Print the script elapsed run time in minutes
    print(f"Time taken: {elapsed_time_minutes:.2f} minutes")
    return {
        'statusCode': 200,
        'body': json.dumps(f"CSV files created successfully as {input_file_key} and {consolidated_file_key}.")
    }


if __name__ == "__main__":
    asyncio.run(main())