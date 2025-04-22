import aiohttp
import asyncio
import configparser
import re
import os
import shutil
import sys
import csv
import time
import datetime
from datetime import datetime
from colorama import Fore
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.views import SheetView, SheetViewList
from openpyxl.utils import get_column_letter
from tqdm.asyncio import tqdm

semaphore = asyncio.Semaphore(64)  # Limit concurrency to 100

# API_CONFIG_PATH = 'API_config-pso.ini'
API_CONFIG_PATH = 'API_config-pcs.ini'
# API_CONFIG_PATH = 'API_config-stm.ini'
# API_CONFIG_PATH = 'API_config-sg.ini'

# include_AccountGroups = False
include_AccountGroups = True # this will generate a lot of data for csv

# Load configuration parameters efficiently
# SSL_VERIFY = config.get("SSL_VERIFY", "False")
SSL_VERIFY = False

start_time = time.time()

def read_api_config():
    config = configparser.ConfigParser()
    config.read(API_CONFIG_PATH)
    baseurl = config.get('URL', 'BaseURL')
    access_key_id = config.get('AUTHENTICATION', 'ACCESS_KEY_ID')
    secret_key = config.get('AUTHENTICATION', 'SECRET_KEY')
    return baseurl, access_key_id, secret_key

class TokenManager:
    def __init__(self, baseurl, access_key_id, secret_key, session):  # Add session parameter
        self.baseurl = baseurl
        self.access_key_id = access_key_id
        self.secret_key = secret_key
        self.token = None
        self.token_expiry_time = 0
        self.lock = asyncio.Lock()
        self.session = session  # Store the session

    async def get_token(self):
        async with self.lock:
            if not self.token or time.time() >= self.token_expiry_time:
                await self._refresh_token()
            return self.token

    async def _refresh_token(self):
        url = f'{self.baseurl}/login'
        headers = {'Content-Type': 'application/json'}
        data = {'username': self.access_key_id, 'password': self.secret_key}
        async with self.session.post(url, headers=headers, json=data, ssl=SSL_VERIFY) as response:  # Use self.session
            if response.status == 200:
                response_data = await response.json()
                self.token = response_data.get('token')
                self.token_expiry_time = time.time() + 480
            else:
                raise Exception(f'Failed to get access token: {await response.text()}')

def sanitize_value(value):
    """Sanitize a value to replace illegal characters for Excel, including adding line breaks."""
    if isinstance(value, str):
        # Replace newline characters (\n) with the Excel line break (\r\n)
        sanitized_value = value.replace('\n', '\r\n')
        # Remove other control characters
        sanitized_value = re.sub(r'[\x00-\x09\x0B\x0C\x0E-\x1F\x7F]', ' ', sanitized_value)
        return sanitized_value.strip()
    return value

# Function to write CSV to a local file
def write_csv_to_file(file_key, data):
    file_path = f"{file_key}"
    with open(file_path, 'w', newline='') as file:
        writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
        writer.writerows(data)
    print(f"CSV file saved to local path {file_path}")


async def make_get_request(url, token_manager, session, max_retries=3, backoff_factor=5):
    """Send a GET request with retry logic and return the JSON response asynchronously."""
    token = await token_manager.get_token()
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    async with semaphore:
        for attempt in range(1, max_retries + 1):
            try:
                async with session.get(url, headers=headers, ssl=SSL_VERIFY) as response:
                    response_text = await response.text()
                    if response.status == 200:
                        return await response.json()
                    else:
                        print(Fore.RED + f"Failed to fetch data {response.status}: {url} {response_text}")
                        print(Fore.WHITE)
                        if attempt == max_retries:
                            return response.status
                        await asyncio.sleep(backoff_factor ** (attempt - 1))
            except aiohttp.ClientError as e:
                print(Fore.RED + f"Error during API call: {e}. URL: {url}, Attempt {attempt} of {max_retries}")
                print(Fore.WHITE)
                if attempt == max_retries:
                    return None
                await asyncio.sleep(backoff_factor ** (attempt - 1))
            except asyncio.TimeoutError:
                print(Fore.RED + f"Request timed out. URL: {url}, Attempt {attempt} of {max_retries}")
                print(Fore.WHITE)
                if attempt == max_retries:
                    return None
                await asyncio.sleep(backoff_factor ** (attempt - 1))
            except Exception as e:
                print(Fore.RED + f"Unexpected error: {e}. URL: {url}, Attempt {attempt} of {max_retries}")
                print(Fore.WHITE)
                if attempt == max_retries:
                    return None
                await asyncio.sleep(backoff_factor ** (attempt - 1))


# Asynchronous function to get policy RQL
async def get_policy_rql(baseurl, token_manager,  search_id, session, retries=3):
    token = await token_manager.get_token()
    if not search_id:
        print(f"Invalid search_id provided: {search_id}")
        return None
    search_url = f"{baseurl}/search/history/{search_id}"
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    for attempt in range(retries):
        try:
            async with session.get(search_url, headers=headers, ssl=SSL_VERIFY, timeout=10) as response:
                response_text = await response.text()
                # Handle a 404 status code by returning 'DELETED'
                if response.status == 404:
                    print(f"Deleted Policy - Policy not found (404) for search_id: {search_id} (attempt {attempt + 1})")
                    # return 'DELETED'
                    return 'Not Available'
                # Process successful response
                if response.status == 200:
                    policy_data = await response.json()
                    if policy_data.get('searchType') == 'asset':
                        policy_rql = policy_data.get('queryWithFindingNames', '')
                    elif policy_data.get('rule', {}).get('type') == 'drift':
                        policy_rql = "{\"category\":\"Drift\",\"resourceTypes\":[]}"
                    else:
                        policy_rql = policy_data.get('query', '')
                    return policy_rql  # Return the policy RQL as a string.
                # Log other status codes and retry
                error_message = f"Error fetching policy RQL from API (attempt {attempt+1}): Status Code: {response.status}"
                print(error_message)
                print(f"Search URL: {search_url}, Search ID: {search_id}")
                print(f"Policy Response: {response_text}")
                # Retry after a delay for non-200 responses
                await asyncio.sleep(2)
        except aiohttp.ClientError as e:
            print(f"ClientError on attempt {attempt+1}: {e}")
            await asyncio.sleep(2)
    # If all retries fail, return
    print(f"All {retries} attempts to fetch policy RQL failed for search_id: {search_id}.")
    return None
    search_data = search_response.json()
    return search_data.get('query', '')


def extract_api_name(rql):
    if not rql:
        return ''
    patterns = {
        "from_api": r"(network|event)\s+from\s+([^\s]+)",
        "api_name": r"api\.name\s*=\s*'([^']*)'",
        "finding_source": r"finding\.source\s*=\s*'([^']*)'",
        "config_network": r"config\s+from\s+network",
        "finding_type": r"finding\.type\s*=\s*'([^']*)'",
        "config_iam": r"config\s+from\s+iam",
        "asset_type": r"asset\.type\s*(?:=\s*'([^']*)'|IN\s*\(\s*'([^']*)'\s*\))",
        "drift": r"Drift:([a-zA-Z]+)",
        "resource_types": r'"resourceTypes":\[(.*?)\]',
        "category": r'"category":"([^"]+)"',
        # "metadata": r"metadata:\s*.*?provider:\s*\"([^\"]*)\".*?category:\s*\"([^\"]*)\"",
        "metadata": r"metadata:\s*.*?provider:\s*([^\s]+).*?category:\s*([^\s]+)",
        # "custom_metadata": r"metadata:\s*.*?category:\s*\"([^\"]+)\".*?provider:\s*\"([^\"]+)\".*?resource_types:\s*-\s*\"([^\"]+)\"",
        "custom_metadata": r"metadata:\s*.*?category:\s*([^\s]+).*?provider:\s*([^\s]+).*?resource_types:\s*-\s*([^\s]+)",
        "terraform_module": r"resource_types:\s*-\s*\"([^\"]+)\".*?attribute:\s*([^\s]+)"
    }
    for key, pattern in patterns.items():
        match re.search(pattern, rql, re.DOTALL):
            case re.Match() as m if key == "from_api":
                return m.group(2)
            case re.Match() as m if key == "api_name":
                return m.group(1)
            case re.Match() as m if key == "finding_source":
                return m.group(1)
            case re.Match() as m if key == "config_network":
                return 'network_exposure'
            case re.Match() as m if key == "finding_type":
                return m.group(1)
            case re.Match() if key == "config_iam":
                return 'iam'
            case re.Match() as m if key == "asset_type":
                return m.group(1) or m.group(2)
            case re.Match() as m if key == "drift":
                return f"IaC:Drift:{m.group(1)}"
            case re.Match() as m if key == "resource_types":
                resource_types = m.group(1).replace('"', '').split(',')
                if resource_types and resource_types[0] == '*':
                    match re.search(patterns["category"], rql):
                        case re.Match() as cm:
                            return process_category(cm.group(1))
                else:
                    api_names = [
                        rt.strip().replace("_", "-").replace("azurerm", "azure")
                        for rt in resource_types
                    ]
                    api_name = ",".join(api_names)
                    match re.search(patterns["category"], rql):
                        case re.Match() as cm:
                            return f"{process_category(cm.group(1))}:{api_name}"
                    return api_name
            case re.Match() as m if key == "metadata":
                provider, category = m.groups()
                return f"{process_category(category)}:{provider}"
            case re.Match() as m if key == "custom_metadata":
                category, provider, resource_type = m.groups()
                return f"{process_category(category)}:{provider}:{resource_type}"
            case re.Match() as m if key == "terraform_module":
                resource_type, attribute = m.groups()
                return f"{resource_type}:{attribute}"

    # If no match is found, return an empty string
    return ''

def process_category(category):
    """
    Function to prepend appropriate prefix based on the category value.
    """
    code_categories = ["Sast", "Secrets", "Licenses"]
    cicd_categories = [
        "System Configuration", "Artifact Integrity Validation", "Credential Hygiene",
        "Data Protection", "Flow Control Mechanisms", "Input Validation",
        "Poisoned Pipeline Execution (PPE)", "Pipeline Flow Control",
        "Pipeline-Based Access Controls (PBAC)", "Supply Chain", "Dependency Chains"
    ]
    if category in code_categories:
        return f"Code:{category}"
    elif category in cicd_categories:
        return f"CI/CD:{category}"
    else:
        return f"IaC:{category}"


def derive_service_name(api):
    """
    Derives the service name from the API name. Handles cases where the input is None or invalid.
    """
    if not api:  # Handle None or empty API value
        return ''
    # Define regex patterns
    patterns = {
        "resource_types": r'"resourceTypes":\[(.*?)\]',
        "category": r'"category":"([^"]+)"'
    }
    # Match resourceTypes and category in the criteria
    match_resource_types = re.search(patterns["resource_types"], api)
    match_category = re.search(patterns["category"], api)
    # Use match-case for evaluating different conditions
    match match_resource_types, api:
        case re.Match() as rt_match, _:
            resource_types = rt_match.group(1).replace('"', '').split(',')
            if resource_types and resource_types[0]:  # If not empty
                service_names = set(derive_service_name(api) for api in resource_types)
                return ",".join(sorted(service_names))
        case None, _ if match_category and '[]' in api:
            return match_category.group(1)
        case None, _ if api.startswith("aws-"):
            parts = api.split("-")
            return "-".join(parts[:2])
        case None, _ if api.startswith("aws-storage-"):
            parts = api.split("-")
            return "-".join(parts[:3]) if len(parts) > 2 else api
        case None, _ if api.startswith("gcloud-cloud-"):
            parts = api.split("-")
            return f"gcloud-cloud-{parts[2]}" if len(parts) > 2 else api
        case None, _ if "api-key" in api:
            parts = api.split("-")
            api_key_index = parts.index('api') + 1
            return '-'.join(parts[:api_key_index + 1]) if len(parts) > api_key_index else api
        case None, _ if api.startswith("gcp-compute") or api.startswith("gcloud-compute"):
            parts = api.split('-')
            if api.startswith(("gcloud-compute-instance", "gcloud-compute-firewall", "gcloud-compute-networks")):
                return '-'.join(parts[:3]) if len(parts) > 2 else api
            return "gcloud-compute"
        case None, _ if any(term in api for term in ["app-engine", "cloud-run", "service-directory", "security-command-center", "block-storage", "object-storage", "app-stream", "file-storage"]):
            parts = api.split('-')
            return '-'.join(parts[:3]) if len(parts) > 2 else api
        case None, _ if api == "azure-kubernetes-cluster":
            return "azure-aks"
        case None, _ if api == "azure-container-instances-container-group":
            return "azure-aci"
        case None, _ if api.startswith("azure-"):
            parts = api.split('-')
            return '-'.join(parts[:3]) if len(parts) > 2 else api
        case None, _ if api.startswith("alibaba-cloud-"):
            parts = api.split("-")
            return "-".join(parts[:3]) if len(parts) > 2 else api
        case _:
            return '-'.join(api.split('-')[:2])


async def get_alerts_data(baseurl, token_manager,  session, status=None):
    token = await token_manager.get_token()
    alerts_url = f"{baseurl}/alert/v1/policy"
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    # Create payload based on status
    payload = {
        "filters": [],
        "fields": ["policyId", "alertCount"],
        "timeRange": {
            "type": "to_now",
            "value": "epoch"
        }
    }
    if status:
        payload["filters"].append({
            "name": "alert.status",
            "value": status,
            "operator": "="
        })
    else:
        payload["filters"].append({
            "name": "alert.status",
            "value": "resolved",
            "operator": "="
        })
    async with session.post(alerts_url, headers=headers, json=payload, ssl=SSL_VERIFY) as response:
        if response.status != 200:
            error_message = f"Error fetching {status if status else 'resolved'} alert counts from API: {await response.text()} (Status Code: {response.status})"
            print(error_message)
            return []
        response_data = await response.json()
        return response_data.get("policies", [])


async def get_all_alerts_count(baseurl, token_manager, session):
    # Fetch all alert counts
    alert_count_data = await get_alerts_data(baseurl, token_manager,  session, status="resolved")
    alert_count_open_data = await get_alerts_data(baseurl, token_manager,  session, status="open")
    alert_count_dismissed_data = await get_alerts_data(baseurl, token_manager,  session, status="dismissed")
    # Create a dictionary with policyId as key and alerts_count, alerts_count_open, and alerts_count_dismissed as values
    alerts_count_dict = {}
    # Populate the dictionary with total alerts count
    for alert in alert_count_data:
        policy_id = alert.get("policyId")
        alerts_count_dict.setdefault(policy_id, {
            "alerts_count": 0,
            "alerts_count_open": 0,
            "alerts_count_dismissed": 0
        })
        alerts_count_dict[policy_id]["alerts_count"] = alert.get("alertCount", 0)
    # Update the dictionary with open alerts count
    for alert in alert_count_open_data:
        policy_id = alert.get("policyId")
        alerts_count_dict.setdefault(policy_id, {
            "alerts_count": 0,
            "alerts_count_open": 0,
            "alerts_count_dismissed": 0
        })
        alerts_count_dict[policy_id]["alerts_count_open"] = alert.get("alertCount", 0)
    # Update the dictionary with dismissed alerts count
    for alert in alert_count_dismissed_data:
        policy_id = alert.get("policyId")
        alerts_count_dict.setdefault(policy_id, {
            "alerts_count": 0,
            "alerts_count_open": 0,
            "alerts_count_dismissed": 0
        })
        alerts_count_dict[policy_id]["alerts_count_dismissed"] = alert.get("alertCount", 0)
    return alerts_count_dict


async def process_compliance_metadata(compliance_metadata, standard_id, customer_name):
    # Handle the case when `compliance_metadata` is empty
    if not compliance_metadata:
        return {
            "stnd_name": '',
            "included": 'NO',
            "requirement_name": '',
            "section_id": '',
            "p_section": ''
        }
    # Step 1: Extract `stnd_name`
    stnd_name = ', '.join(
        metadata.get('standardName', '') for metadata in compliance_metadata
    )
    # Step 2: Initialize default values
    included = "NO"
    requirement_name = ''
    section_id = ''
    p_section = ''
    # Step 3: Find entry with matching `standardId`
    matching_entry = None
    custom_assigned_entry = None
    for metadata in compliance_metadata:
        # Check if this entry has a matching `standardId`
        if metadata.get('standardId') == standard_id:
            matching_entry = metadata
            break  # Stop searching if we find a direct match
        # Track the last customAssigned entry as a fallback
        if metadata.get('customAssigned'):
            custom_assigned_entry = metadata
    # Step 4: Assign values based on the found entry
    if matching_entry:
        # Found a direct match with `standardId`
        included = customer_name
        stnd_name = metadata.get('standardName', '')
        requirement_name = matching_entry.get('requirementId', '')
        section_id = matching_entry.get('sectionId', '')
        p_section = matching_entry.get('sectionId', '')
    elif custom_assigned_entry:
        # No direct match, use the last customAssigned entry
        requirement_name = custom_assigned_entry.get('requirementId', '')
        section_id = custom_assigned_entry.get('sectionId', '')
    elif compliance_metadata:
        # No `customAssigned` entry, use the last entry in the list
        last_entry = compliance_metadata[-1]
        requirement_name = last_entry.get('requirementId', '')
        section_id = last_entry.get('sectionId', '')
    # Step 5: Return processed values as a dictionary
    return {
        "stnd_name": stnd_name,
        "included": included,
        "requirement_name": requirement_name,
        "section_id": section_id,
        "p_section": p_section
    }


def handle_build_children(children, is_custom=False):
    """
    Extract data from build children and return RQL, Checkov ID, and recommendation.
    For custom policies, if 'criteria' is missing, it returns data from 'metadata'.
    """
    for child in children:
        # Handle custom policies with 'metadata' instead of 'criteria'
        if is_custom and not child.get('criteria', ''):
            metadata = child.get('metadata', {})
            return metadata.get('code', ''), metadata.get('checkovId', ''), child.get('recommendation', '')

        # Handle standard policies with 'criteria'
        return child.get('criteria', ''), child.get('metadata', {}).get('checkovId', ''), child.get('recommendation', '')
    # Default return if no matching child found
    return '', '', ''


# Define a function to handle build and run policies
async def handle_build_and_run_policy(policy_data, baseurl, token_manager, session):
    search_id = policy_data.get('rule', {}).get('criteria', '')  # Run criteria
    if not search_id and policy_data.get('rule', {}).get('type') == 'drift':
        # print(f"Problem search_id - {policy_data}")
        run_policy_rql = "{\"category\":\"Drift\",\"resourceTypes\":[]}"
    else:
        # Fetch run RQL using search_id
        run_policy_rql = await get_policy_rql(baseurl, token_manager, search_id, session)
    checkov_id, build_policy_rql, build_recommendation = None, '', ''
    recommendation = policy_data.get('recommendation', '')

    # Process build part
    children = policy_data.get('rule', {}).get('children', [])
    for child in children:
        if child.get('type') == 'build':
            # Extract build-related values
            build_policy_rql = child.get('criteria', '')
            checkov_id = child.get('metadata', {}).get('checkovId', '')
            build_recommendation = child.get('recommendation', '')
            break
    # Combine run and build RQL
    combined_policy_rql = f"{run_policy_rql}\r\n{build_policy_rql}".strip()
    # Extract API name and service name based on the combined RQL
    api_name = extract_api_name(build_policy_rql if not run_policy_rql else run_policy_rql)
    service_name = derive_service_name(api_name)
    # Prepare recommendation combining both run and build parts
    combined_recommendation = f"{recommendation} \r\nBuild Recommendation: {build_recommendation}\r\n\r\nIaC: {build_recommendation}".strip()
    # Return the relevant fields for further processing or storing in the final output
    return run_policy_rql, combined_policy_rql, api_name, service_name, checkov_id, combined_recommendation

async def fetch_policy_data(baseurl, policy_id, customer_name, standard_id, session, token_manager, alerts_count_dict, alerts_rules_dict):

    get_policy_info_url = f"{baseurl}/policy/{policy_id}"
    # policy_data = await get_policy_info(baseurl, token_manager,  policy_id, session)
    policy_data = await make_get_request(get_policy_info_url, token_manager, session) or {}
    if not policy_data:
        print(f"Unable to fetch policy data: {policy_id}")
        return None
    # Extract main fields from the policy data
    system_default = policy_data.get('systemDefault', '')
    cloud_type = policy_data.get('cloudType', '')
    labels = ",".join(policy_data.get('labels', [])) if policy_data.get('labels') else ''
    policy_mode = policy_data.get('policyMode', '')
    policy_id = policy_data.get('policyId', '')
    policy_type = policy_data.get('policyType', '')
    # policy_subtype = policy_data.get('policySubTypes', [''])[0]
    policy_subtypes = policy_data.get('policySubTypes', [])
    policy_subtypes_str = ",".join(policy_subtypes) if policy_subtypes else ''
    policy_name = policy_data.get('name', '')
    description = policy_data.get('description', '')
    severity = policy_data.get('severity', '')
    policy_upi = policy_data.get('policyUpi', '')
    last_modified_by = policy_data.get('lastModifiedBy', '')
    recommendation = policy_data.get('recommendation', '')
    remediation_value = policy_data.get('remediation', '')
    remediation = str(remediation_value).replace('{', '').replace('}', '').replace(', ', ',\r\n')
    remediable = policy_data.get('remediable', '')
    last_modified_on = policy_data.get('lastModifiedOn', 0)
    # Extract `complianceMetadata` from the response
    compliance_metadata = policy_data.get("complianceMetadata", [])
    # Process compliance metadata asynchronously
    metadata_info = await process_compliance_metadata(compliance_metadata, standard_id, customer_name)
    # Extract values from `metadata_info`
    stnd_name = metadata_info["stnd_name"]
    included = metadata_info["included"]
    requirement_name = metadata_info["requirement_name"]
    section_id = metadata_info["section_id"]
    if included == 'customer_name':
        p_section = section_id
    else:
        p_section = metadata_info["p_section"]
    enabled = policy_data.get('enabled', False)
    alerts_count = alerts_count_open = alerts_count_dismissed = ''
    if enabled:
        if alerts_data := alerts_count_dict.get(policy_id):
            alerts_count = alerts_data.get("alerts_count", 0)
            alerts_count_open = alerts_data.get("alerts_count_open", 0)
            alerts_count_dismissed = alerts_data.get("alerts_count_dismissed", 0)
        # Update the `included` status based on alert counts if they are integers.
        if included == 'NO' and isinstance(alerts_count_dismissed, int) and alerts_count_dismissed > 0:
            included = 'REVIEW'
        elif included == 'NO' and isinstance(alerts_count_open, int) and alerts_count_open > 0:
            included = 'MAYBE'
    # Convert timestamp to a readable datetime format
    last_modified_on_str = datetime.fromtimestamp(last_modified_on / 1000).strftime('%Y-%m-%d %H:%M:%S') if last_modified_on else ''
    policyCategory = policy_data.get('policyCategory', '')
    findingTypes = ",".join(str(findingType) for findingType in policy_data.get('findingTypes', []) if findingType) if policy_data.get('findingTypes') else ''
    policyClass = policy_data.get('policyClass', '')
    owner = policy_data.get('owner', '')
    search_id = policy_data.get('rule', {}).get('criteria', '')
    # checkov_id, build_policy_rql, build_recommendation, api_name, service_name = None, '', '', '', ''
    checkov_id, build_recommendation, api_name, service_name = None, '', '', ''
    # Map certain policy types to specific API names and service names
    type_service_map = {
        'iam': (f"{policy_type}", f"{policy_subtypes_str}_{policy_type}"),
        'audit_event': (f"{policy_type}", f"{policy_type}"),
        'anomaly': (f"{policy_type}_{policy_subtypes_str}", f"{search_id}"),
        'api': ('network_event_api', f"{policy_subtypes_str}_{policy_type}"),
        'malware': (f"{policy_type}", f"{policy_type}_{policy_subtypes_str}"),
        'grayware': (f"{policy_type}", f"{policy_type}_{policy_subtypes_str}"),
        'data': (f"{policy_type}", f"{policy_subtypes_str}_{policy_type}"),
        'workload_vulnerability': (f"{policy_type}", f"{policy_type}"),
        'workload_incident': (f"{policy_type}", f"{policy_type}")
    }
    # Process common RQL handling for certain policy subtypes
    policy_subtypes_to_rql = ['run', 'audit', 'network_event', 'event', 'network_exposure', 'network_config', 'misconfig', 'misconfig_and_event', 'permissions']
    # Handle build policies
    if policy_subtypes_str == 'build':
        children = policy_data.get('rule', {}).get('children', [])
        policy_rql, checkov_id, build_recommendation = handle_build_children(
            children, is_custom=(policy_data.get('policyMode') == 'custom'))
        recommendation = build_recommendation
        if checkov_id:
            policy_upi = checkov_id
        api_name = extract_api_name(policy_rql)
        service_name = derive_service_name(api_name)
    # Handle run_and_build policies
    elif policy_subtypes_str == 'build,run':
        run_policy_rql, policy_rql, api_name, service_name, checkov_id, recommendation = await handle_build_and_run_policy(policy_data, baseurl, token_manager,  session)
        # Update Policy UPI with Checkov ID if applicable
        if checkov_id and checkov_id not in policy_upi:
            policy_upi = f"{policy_upi}-{checkov_id}" if policy_upi else checkov_id
    elif search_id and policy_subtypes_str in policy_subtypes_to_rql:
        # Fetch the RQL for the specific subtype
        policy_rql = await get_policy_rql(baseurl, token_manager,  search_id, session)
        # Extract API name and service name if RQL exists
        api_name = extract_api_name(policy_rql)
        service_name = derive_service_name(api_name)
    else:
        # Fallback to use the search_id if no specific RQL is found
        policy_rql = search_id
        # Ensure the default return is a tuple
        api_name, service_name = type_service_map.get(policy_type, (api_name, service_name))
    if policy_rql == 'DELETED':
        included = 'DELETED'
    # Handle specific policies that don't support standards
    if policy_type in ['attack_path', 'iam', 'api', 'workload_vulnerability', 'workload_incident'] or policy_subtypes_str in ["network_config"]:
        # p_requirement = "No Support for Compliance Standard"
        included = "NO_SUPPORT"
    alert_rules = await get_alert_rule_names_for_policy(policy_id, cloud_type, severity, alerts_rules_dict)
    if include_AccountGroups:
        AccountGroups = await get_AccountGroups_for_policy(policy_id, alerts_rules_dict)
    else:
        AccountGroups = ''
    # Prepare the new row data
    reason = f"Discovered {datetime.now().strftime('%Y/%m/%d/%H:%M')}"
    # For reference ONLY here - headers = [
    # "Standard", "Included", "Resolved", "Open", "Dismissed", "Enabled", "Rules", "CloudType", "Service", "PolicyName", "FindingTypes", "Severity", "Owner",
    # "PolicyType", "PolicySubTypes", "Requirement", "Section", "SectionOption", "Labels", "Remediable",
    # "PolicyClass", "API Name", "Description", "Recommendation", "Remediation", "Policy RQL", "LastModifiedBy",
    # "LastModifiedOn", "Policy UPI", "SystemDefault", "PolicyMode", "PolicyCategory", "AccountGroups", "PolicyID", "search_id", "savedSearch", "Reason"
    # ]
    return [
        stnd_name, included, p_section, alerts_count, alerts_count_open, alerts_count_dismissed, enabled, alert_rules, cloud_type, service_name, policy_name, findingTypes, severity, owner,
        policy_type, policy_subtypes_str, requirement_name, section_id, labels, remediable, policyClass, api_name, description, recommendation, remediation,
        policy_rql, last_modified_by, last_modified_on_str, policy_upi, system_default, policy_mode, policyCategory, AccountGroups, policy_id, search_id, "true" if search_id else "false", reason
    ]


async def process_policies_for_cloud_type(cloud_type, baseurl, token_manager, customer_name, standard_id, alerts_count_dict, alerts_rules_dict, session):
    # Fetch policy IDs for the given cloud type
    policy_url = f"{baseurl}/policy?cloud.type={cloud_type}"
    policy_data = await make_get_request(policy_url, token_manager, session)
    policy_ids = [policy['policyId'] for policy in policy_data]
    print(f"Fetched {len(policy_ids)} policies for cloud type {cloud_type}.")
    if not policy_ids:
        print(f"No policies found for cloud type {cloud_type}.")
        return {}
    # Prepare tasks for fetching policy data
    tasks = [
        fetch_policy_data(
            baseurl=baseurl,
            policy_id=policy_id,
            customer_name=customer_name,
            standard_id=standard_id,
            session=session,
            token_manager=token_manager,
            alerts_count_dict=alerts_count_dict,
            alerts_rules_dict=alerts_rules_dict
    )
        for policy_id in policy_ids
    ]
    # Gather results asynchronously with error handling
    results = await asyncio.gather(*tasks, return_exceptions=True)
    # Create a dictionary to store unique results keyed by `PolicyID`
    policy_data_dict = {}
    for idx, result in enumerate(results):
        # Retrieve the policy ID using the index from policy_ids
        policy_id = policy_ids[idx] if idx < len(policy_ids) else "Unknown"
        if isinstance(result, Exception):
            # Log more details including the policy ID and exception message.
            print(f"Error fetching policy data for policy_id {policy_id}: {result}")
        elif result:
            # Check if the `PolicyID` already exists in the dictionary.
            if policy_id in policy_data_dict:
                # If it exists, you can add logic here to decide which policy to keep.
                # For example, prefer the first occurrence or based on a specific cloud_type or condition.
                print(f"Duplicate policy_id {policy_id} detected across cloud types; keeping the first occurrence.")
            else:
                # Store the result only if the PolicyID is not already in the dictionary.
                policy_data_dict[policy_id] = result
    # print(f"Cloud Type: {cloud_type}, Number of Unique Policies Processed: {len(policy_data_dict)}")
    # Return the dictionary of policy data keyed by `PolicyID`
    return policy_data_dict


# async def process_labels(rows, headers, standard_label, customer_name):
#     updated_rows = []
#     customer_ID = customer_name[:2].upper()  # Get the first two letters of Customer name in capitals
#     labels_index = headers.index("Labels")
#     policy_mode_index = headers.index("PolicyMode")
#     cloud_type_index = headers.index("CloudType")
#     p_section_index = headers.index("SectionOption")
#     for row in rows:
#         labels = row[labels_index]
#         policy_mode = row[policy_mode_index]
#         cloud_type = row[cloud_type_index].upper()
#         p_section = row[p_section_index]
#         # Ensure p_section is treated as a string, even if it's None or NaN
#         p_section = str(p_section)
#         # Split labels and filter out any that start with {customer_ID}_
#         labels_list = [label for label in labels.split(',') if label and not label.startswith(f'{customer_ID}_')]
#         if standard_label not in labels_list:
#             labels_list.append(standard_label)
#         if p_section and f'{customer_ID}_{cloud_type}_{p_section}' not in labels_list:
#             labels_list.append(f'{customer_ID}_{cloud_type}_{p_section}')
#         # Add Custom for rows that don't equal to LastModifiedBy = "Prisma Cloud System Admin" and don't already have Custom
#         if policy_mode == "custom" and 'Custom' not in labels_list:
#             labels_list.append('Custom')
#         # Update the Labels column with the filtered and updated list
#         row[labels_index] = ','.join(labels_list)
#         updated_rows.append(row)
#     return updated_rows


async def get_policies_alert_rules(baseurl, token_manager, session):
    token = await token_manager.get_token()
    alerts_rule_url = f"{baseurl}/v2/alert/rule"
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    async with session.get(alerts_rule_url, headers=headers, ssl=SSL_VERIFY) as response:
        if response.status != 200:
            error_message = f"Error fetching alert rules from API: {await response.text()} (Status Code: {response.status})"
            print(error_message)
            return {}
        alerts_rule_data = await response.json()
    account_groups_url = f"{baseurl}/cloud/group"
    account_groups = await make_get_request(account_groups_url, token_manager, session)
    alerts_rules_dict = {}
    for rule in alerts_rule_data:
        rule_name = rule.get("name", "")
        scan_all = rule.get("scanAll", False)
        target_resource_list = rule.get("target", {}).get("targetResourceList", {})
        alert_rule_policy_filter = rule.get("alertRulePolicyFilter", {})
        # Append (auto_dismiss) only if action is AUTO_DISMISS, enabled is True, and ids list is not empty
        if (
            target_resource_list.get("action") == "AUTO_DISMISS" and
            target_resource_list.get("enabled") and
            isinstance(target_resource_list.get("ids"), list) and
            len(target_resource_list.get("ids")) > 0
        ):
            rule_name = f"{rule_name} (auto_dismiss)"
            print(f"adding auto_dismiss to {rule_name}")
        # Handle "scanAll" rules
        if scan_all:
            alerts_rules_dict[rule_name] = {
                "policies": "all",
                "notificationChannels": rule.get("notificationChannels", []),
                "accountGroupsName": [
                    group['name'] for group in account_groups
                    if group['id'] in rule.get("target", {}).get("accountGroups", [])
                ],
                "resourceLists": target_resource_list.get("ids", []),
                "alertRulePolicyFilter": alert_rule_policy_filter
            }
        else:
            policy_ids = rule.get("policies", [])
            # Create the alert rule entry in the dictionary for specific policies
            alerts_rules_dict[rule_name] = {
                "policies": policy_ids,
                "notificationChannels": rule.get("notificationChannels", []),
                "accountGroupsName": [
                    group['name'] for group in account_groups
                    if group['id'] in rule.get("target", {}).get("accountGroups", [])
                ],
                "resourceLists": target_resource_list.get("ids", []),
                "alertRulePolicyFilter": alert_rule_policy_filter
            }
    return alerts_rules_dict


async def get_alert_rule_names_for_policy(policy_id, cloud_type, severity, alerts_rules_dict):
    matching_rules = []
    for rule_name, rule_details in alerts_rules_dict.items():
        policies = rule_details.get("policies", [])
        alert_rule_filter = rule_details.get("alertRulePolicyFilter", {})
        # Check if policy id matches, rule is "scanAll", or filter matches severity and cloud type
        if (
            policies == "all" or
            policy_id in policies or
            (
                severity in alert_rule_filter.get("policy.severity", []) and
                cloud_type in alert_rule_filter.get("cloud.type", [])
            )
        ):
            matching_rules.append(rule_name)
            # Log matching rules with auto_dismiss for clarity
            if "auto_dismiss" in rule_name:
                print(f"Matching rule with auto_dismiss: {rule_name}")
    matching_rules.sort()
    return ", ".join(matching_rules)


async def get_AccountGroups_for_policy(policy_id, alerts_rules_dict):
    matching_rules = []
    for rule_name, rule_details in alerts_rules_dict.items():
        if rule_details["policies"] == "all" or policy_id in rule_details["policies"]:
            notification_channels = rule_details.get("notificationChannels", [])
            account_groups = rule_details.get("accountGroupsName", [])
            details = f"{rule_name} - Notifications:{','.join(notification_channels)}, AccountGroups:{','.join(account_groups)};"
            matching_rules.append(details)
    matching_rules.sort()
    return ", ".join(matching_rules)


async def prompt_for_standard_selection(standards_data):
    """Prompt the user to select a compliance standard, sorted by policiesAssignedCount, and return the selected standard details."""
    # Sort standards by policiesAssignedCount in descending order
    sorted_standards = sorted(standards_data, key=lambda x: x.get("name", 0), reverse=False)
    print("Available Compliance Standards (sorted by policies assigned):")
    for index, standard in enumerate(sorted_standards, start=1):
        print(Fore.GREEN + f"{index}. {standard.get('name')}" + Fore.WHITE + f" - Policies Assigned: {standard.get('policiesAssignedCount')}, "
              f"cloudType: {', '.join(standard.get('cloudType', []))}, createdBy: {standard.get('createdBy')}")
    # Prompt the user for a selection
    try:
        selection = int(input("Select a compliance standard by number: ")) - 1
        if selection < 0 or selection >= len(sorted_standards):
            print("Invalid selection. Please try again.")
            return await prompt_for_standard_selection(standards_data)  # Retry if invalid
        selected_standard = sorted_standards[selection]
        # Extract the cloud types from the selected standard
        stnd_id = selected_standard.get("id")
        cloud_types = selected_standard.get("cloudType", [])
        print(f"Selected Standard ID: {stnd_id}")
        print(f"Cloud Types for this standard: {cloud_types}")
        return selected_standard
    except ValueError:
        print("Invalid input. Please enter a number.")
        return await prompt_for_standard_selection(standards_data)


async def main():
    test_event = {
        # "Customer": "SocGen",
        "Customer": "Awesome Inc",
        # "Customer": "Widget",
        "bucket_name": "data",
        "filename": "compliance_all_policies",
        # "standard_id": "a0ea1077-424f-45fd-994e-4caef6d4d9de", # AWS Foundational Security Best Practices standard
        "standard_id": "581cc66e-4559-4cc6-be12-470d5d34077b", # PCS Awesome Standard
        # "standard_id": "3b3c3f52-3cdf-4849-8acf-09276cc8e0d8", # PSO MTSBv2_AWS Standard
        "secret_name": "compliance-policy"
    }
    baseurl, access_key_id, secret_key = read_api_config()
    async with aiohttp.ClientSession() as session:

        token_manager = TokenManager(baseurl, access_key_id, secret_key, session)
        # standard_id = test_event.get('standard_id', None)
        customer_name = test_event.get('Customer', 'Customer_Test')
        customer_label = customer_name.replace(' ', '_')
        bucket_name = test_event.get('bucket_name', 'pc-compliance-pl')
        filename = test_event.get('filename', 'compliance_policies')

        headers = [
            "Standard", "Included", "SectionOption", "Resolved", "Open", "Dismissed", "Enabled", "Rules", "CloudType", "Service",
            "PolicyName", "FindingTypes", "Severity", "Owner", "PolicyType", "PolicySubTypes",
            "Requirement", "Section",  "Labels", "Remediable", "PolicyClass", "API Name",
            "Description", "Recommendation", "Remediation", "Policy RQL", "LastModifiedBy", "LastModifiedOn",
            "Policy UPI", "SystemDefault", "PolicyMode", "PolicyCategory", "AccountGroups", "PolicyID",
            "search_id", "savedSearch", "Reason"
        ]
        standards_url = f"{baseurl}/compliance"
        standards_data = await make_get_request(standards_url, token_manager, session)
        filtered_standards_data = [standard for standard in standards_data if not standard.get("systemDefault", True)]
        # Prompt the user to select a compliance standard
        selected_standard = await prompt_for_standard_selection(filtered_standards_data)
        if not selected_standard:
            print("No compliance standard selected.")
            return
        # Extract the stnd_id and cloud_types from the selected standard
        standard_id = selected_standard.get("id")
        cloud_types = selected_standard.get("cloudType", [])
        stnd_name = selected_standard.get('name')
        stnd_label = stnd_name.replace(' ', '_')
        backup_file_key = f"{bucket_name}/{filename}_{customer_label}_{stnd_label}_latest_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
        updated_file_key = f"{bucket_name}/{filename}_{customer_label}_{stnd_label}_latest.xlsx"
        # Check if the file exists
        if os.path.exists(updated_file_key):
            user_input = input(f"File {updated_file_key} already exists. Do you want to override it? (yes/no): ").strip().lower()
            if user_input == "yes":
                print(f"Overriding the file: {updated_file_key}")
                try:
                    # Create a backup
                    shutil.copy(updated_file_key, backup_file_key)
                    print(f"Backup created: {backup_file_key}")
                except Exception as e:
                    print(f"Failed to create a backup: {str(e)}")
            elif user_input == "no":
                print("Exiting script as per user choice.")
                sys.exit(0)
            else:
                print("Invalid input. Exiting script.")
                sys.exit(1)
        print(f"Selected Standard {stnd_name} - ID: {standard_id}")
        print(f"Cloud Types for this standard: {cloud_types}")
        print(f"Collecting {customer_name} policies for {stnd_name}")
        # Create a workbook and add headers to the Excel sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"{stnd_name} Policies"
        # # Fetch requirements asynchronously
        requirements_url = f"{baseurl}/compliance/{standard_id}/requirement"
        requirements_data = await make_get_request(requirements_url, token_manager, session)
        # Process each requirement and fetch associated sections
        p_section_options = []
        for requirement in requirements_data:
            requirement_id = requirement.get('id')
            if not requirement_id:
                continue
            # Fetch sections for each requirement
            sections_url = f"{baseurl}/compliance/{requirement_id}/section"
            sections_data = await make_get_request(sections_url, token_manager, session)
            # Extract section IDs and append to the list
            for section in sections_data:
                section_id = section.get('sectionId')
                if section_id:
                    p_section_options.append(section_id)
        alerts_count_dict = await get_all_alerts_count(baseurl, token_manager, session)
        alerts_rules_dict = await get_policies_alert_rules(baseurl, token_manager, session)
        # cloud_types = ["aws", "azure", "gcp", "oci", "alibaba_cloud", "ibm"] extracted from the selected Compliance standard
        # Define the shading color for dropdown columns (light yellow in this example)
        dropdown_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        # Write headers to the Excel file
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            # Define dropdown options for "Included" and "SectionOption"
        included_options = ["YES", "NO", "REVIEW", "MAYBE", "REMOVE"]
        # Create DataValidation object for "Included"
        included_validation = DataValidation(type="list", formula1=f'"{",".join(included_options)}"')
        included_validation.error = 'Invalid value entered. Please select a value from the dropdown list.'
        included_validation.errorTitle = 'Invalid Entry'
        # Create DataValidation object for "SectionOption"
        p_section_validation = DataValidation(type="list", formula1=f'"{",".join(p_section_options)}"')
        p_section_validation.error = 'Invalid value entered. Please select a value from the dropdown list.'
        p_section_validation.errorTitle = 'Invalid Entry'
        # Store all results in a dictionary to ensure uniqueness by `PolicyID`
        all_results_dict = {}
        # Process all cloud types and gather results
        for cloud_type in tqdm(cloud_types, desc=f"Retrieving policies details"):
            policy_data_dict = await process_policies_for_cloud_type(cloud_type, baseurl, token_manager, customer_name, standard_id, alerts_count_dict, alerts_rules_dict, session)
            # Merge into `all_results_dict` to ensure unique `PolicyID`s
            all_results_dict.update(policy_data_dict)
        # Convert the dictionary values back to a list for further processing
        all_results = list(all_results_dict.values())
        print(f"Total number of policies processed: {len(all_results)}")
        # Filter `all_results` for policies with `PolicyMode == "custom"`
        custom_results = [result for result in all_results if result[headers.index("PolicyMode")] == "custom"]
        # Filter `all_results` where provided standard in `Included` equals `customer_name`
        included_results = [result for result in all_results if result[headers.index("Included")] == customer_name]
        # Call `process_labels` with the filtered `included_results`
        # if included_results:
        #     included_results = await process_labels(included_results, headers, stnd_label, customer_name)
        # Combine `custom_results` and `included_results` back into `all_results` before deduplication
        combined_results_dict = {result[headers.index("PolicyID")]: result for result in (custom_results + included_results)}
        # Add remaining non-custom, non-included results to `combined_results_dict`
        for result in all_results:
            policy_id = result[headers.index("PolicyID")]
            if policy_id not in combined_results_dict:
                combined_results_dict[policy_id] = result
        # Convert the combined results back to a list for sorting
        combined_results = list(combined_results_dict.values())
        # Sort `combined_results` such that included results come first, each sorted by API Name
        included_index = headers.index("Included")
        api_name_index = headers.index("API Name")
        sorted_policies = sorted(
            combined_results,
            key=lambda x: (0 if x[included_index] == customer_name else 1, x[api_name_index])
        )
        # Write the sorted results to the Excel sheet, adding `stnd_name` for each row in the "Standard" column
        print(f"Total number of unique policies processed: {len(sorted_policies)}")

        # Determine the last row for applying validation and shading
        data_end_row = len(sorted_policies) + 1  # Assuming headers are in row 1

        # Apply validations to the entire data range in the columns
        included_col_letter = openpyxl.utils.get_column_letter(2)  # Column B
        p_section_col_letter = openpyxl.utils.get_column_letter(3)  # Column C
        included_validation.add(f"{included_col_letter}2:{included_col_letter}{data_end_row}")
        p_section_validation.add(f"{p_section_col_letter}2:{p_section_col_letter}{data_end_row}")
        sheet.add_data_validation(included_validation)
        sheet.add_data_validation(p_section_validation)

        # Apply shading to the cells (starting from row 2)
        for row_num in range(2, data_end_row + 1):
            included_cell = sheet.cell(row=row_num, column=2)  # Column B for "Included"
            p_section_cell = sheet.cell(row=row_num, column=3)  # Column C for "SectionOption"
            # Apply shading
            included_cell.fill = dropdown_fill
            p_section_cell.fill = dropdown_fill
        # Write the sorted results to the Excel sheet
        for row_num, policy_data in enumerate(sorted_policies, 2):
            for col_num, cell_value in enumerate(policy_data, 1):
                sanitized_value = sanitize_value(cell_value)
                sheet.cell(row=row_num, column=col_num, value=sanitized_value)
        sheet.calculate_dimension()
        sheet.views = SheetViewList(sheetView=[SheetView(tabSelected=True, workbookViewId=0)])
        # Save the workbook
        workbook.save(updated_file_key)
        print(f"All data written to {updated_file_key} successfully.")
        # Record the end time
        end_time = time.time()
        # Calculate the elapsed time
        elapsed_time_seconds = end_time - start_time
        # Convert elapsed time to minutes
        elapsed_time_minutes = elapsed_time_seconds / 60
        # Print the script elapsed run time in minutes
        print(f"Time taken: {elapsed_time_minutes:.2f} minutes")

if __name__ == "__main__":
    asyncio.run(main())
