import json
import csv
import requests
import aiohttp
import asyncio
import configparser
import shutil

from io import StringIO
from collections import defaultdict
import re
# import threading
import time
import datetime
from datetime import datetime, timedelta
from colorama import Fore
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from tqdm.asyncio import tqdm

semaphore = asyncio.Semaphore(64)  # Limit concurrency to 100

# API_CONFIG_PATH = 'API_config-pso.ini'
API_CONFIG_PATH = 'API_config-pcs.ini'
# API_CONFIG_PATH = 'API_config-stm.ini'
# API_CONFIG_PATH = 'API_config-sg.ini'

# include_AccountGroups = False
include_AccountGroups = True # this will generate a lot of data for csv

# Load configuration parameters efficiently
# SSL_VERIFY = config.get("SSL_VERIFY", "False")
SSL_VERIFY = False

start_time = time.time()

def read_api_config():
    config = configparser.ConfigParser()
    config.read(API_CONFIG_PATH)
    baseurl = config.get('URL', 'BaseURL')
    access_key_id = config.get('AUTHENTICATION', 'ACCESS_KEY_ID')
    secret_key = config.get('AUTHENTICATION', 'SECRET_KEY')
    # print(f"{baseurl} {access_key_id} {secret_key}")
    return baseurl, access_key_id, secret_key


class TokenManager:
    def __init__(self, baseurl, access_key_id, secret_key, session):  # Add session parameter
        self.baseurl = baseurl
        self.access_key_id = access_key_id
        self.secret_key = secret_key
        self.token = None
        self.token_expiry_time = 0
        self.lock = asyncio.Lock()
        self.session = session  # Store the session

    async def get_token(self):
        async with self.lock:
            if not self.token or time.time() >= self.token_expiry_time:
                await self._refresh_token()
            return self.token

    async def _refresh_token(self):
        url = f'{self.baseurl}/login'
        headers = {'Content-Type': 'application/json'}
        data = {'username': self.access_key_id, 'password': self.secret_key}
        async with self.session.post(url, headers=headers, json=data, ssl=SSL_VERIFY) as response:  # Use self.session
            if response.status == 200:
                response_data = await response.json()
                self.token = response_data.get('token')
                self.token_expiry_time = time.time() + 480
            else:
                raise Exception(f'Failed to get access token: {await response.text()}')

def sanitize_value(value):
    """Sanitize a value to replace illegal characters for Excel, including adding line breaks."""
    if isinstance(value, str):
        # Replace newline characters (\n) with the Excel line break (\r\n)
        sanitized_value = value.replace('\n', '\r\n')
        # Remove other control characters
        sanitized_value = re.sub(r'[\x00-\x09\x0B\x0C\x0E-\x1F\x7F]', ' ', sanitized_value)
        return sanitized_value.strip()
    return value


async def make_get_request(url, token_manager, session, max_retries=3, backoff_factor=5):
    """Send a GET request with retry logic and return the JSON response asynchronously."""
    token = await token_manager.get_token()
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    async with semaphore:
        for attempt in range(1, max_retries + 1):
            try:
                async with session.get(url, headers=headers, ssl=SSL_VERIFY) as response:
                    response_text = await response.text()
                    if response.status == 200:
                        return await response.json()
                    else:
                        print(Fore.RED + f"Failed to fetch data {response.status}: {url} {response_text}")
                        print(Fore.WHITE)
                        if attempt == max_retries:
                            return response.status
                        await asyncio.sleep(backoff_factor ** (attempt - 1))
            except aiohttp.ClientError as e:
                print(Fore.RED + f"Error during API call: {e}. URL: {url}, Attempt {attempt} of {max_retries}")
                print(Fore.WHITE)
                if attempt == max_retries:
                    return None
                await asyncio.sleep(backoff_factor ** (attempt - 1))
            except asyncio.TimeoutError:
                print(Fore.RED + f"Request timed out. URL: {url}, Attempt {attempt} of {max_retries}")
                print(Fore.WHITE)
                if attempt == max_retries:
                    return None
                await asyncio.sleep(backoff_factor ** (attempt - 1))
            except Exception as e:
                print(Fore.RED + f"Unexpected error: {e}. URL: {url}, Attempt {attempt} of {max_retries}")
                print(Fore.WHITE)
                if attempt == max_retries:
                    return None
                await asyncio.sleep(backoff_factor ** (attempt - 1))


# Function to write CSV to a local file
def write_csv_to_file(file_key, data):
    file_path = f"{file_key}"
    with open(file_path, 'w', newline='') as file:
        writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
        writer.writerows(data)
    print(f"CSV file saved to local path {file_path}")


# Function to read Excel data from a local file
def read_excel_from_file(folder, file_key):
    file_path = f"{folder}/{file_key}"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data


# Function to write Excel to a local file
def write_excel_to_file(folder, file_key, data):
    file_path = f"{folder}/{file_key}"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row_num, row in enumerate(data, start=1):
        for col_num, cell_value in enumerate(row, start=1):
            sanitized_value = sanitize_value(cell_value)  # Assuming you have a sanitize_value function
            sheet.cell(row=row_num, column=col_num, value=sanitized_value)
    workbook.save(file_path)
    print(f"Excel file saved to local path {file_path}")


async def prompt_for_standard_selection(standards_data):
    """Prompt the user to select a compliance standard, sorted by policiesAssignedCount, and return the selected standard details."""
    # Sort standards by policiesAssignedCount in descending order
    sorted_standards = sorted(standards_data, key=lambda x: x.get("name", 0), reverse=False)
    print("Available Compliance Standards (sorted by policies assigned):")
    for index, standard in enumerate(sorted_standards, start=1):
        print(Fore.GREEN + f"{index}. {standard.get('name')}" + Fore.WHITE + f" - Policies Assigned: {standard.get('policiesAssignedCount')}, "
              f"cloudType: {', '.join(standard.get('cloudType', []))}, createdBy: {standard.get('createdBy')}")
    # Prompt the user for a selection
    try:
        selection = int(input("Select a compliance standard by number: ")) - 1
        if selection < 0 or selection >= len(sorted_standards):
            print("Invalid selection. Please try again.")
            return await prompt_for_standard_selection(standards_data)  # Retry if invalid
        selected_standard = sorted_standards[selection]
        # Extract the cloud types from the selected standard
        stnd_id = selected_standard.get("id")
        cloud_types = selected_standard.get("cloudType", [])
        print(f"Selected Standard ID: {stnd_id}")
        print(f"Cloud Types for this standard: {cloud_types}")
        return selected_standard
    except ValueError:
        print("Invalid input. Please enter a number.")
        return await prompt_for_standard_selection(standards_data)


async def generate_compliance_summary(baseurl, token_manager, standard_id, session):
    # Dictionary to store the summary data
    summary_data = [["Section ID", "Section Description", "Policies Assigned Count", "Requirement Name", "Requirement Description"]]
    requirement_url = f"{baseurl}/compliance/{standard_id}/requirement"
    requirements_data = await make_get_request(requirement_url, token_manager, session)

    if not requirements_data:
        print("No requirements found for the standard.")
        return summary_data

    for requirement in requirements_data:
        requirement_id = requirement.get("id")
        requirement_name = requirement.get("name")
        requirement_description = requirement.get("description")

        section_url = f"{baseurl}/compliance/{requirement_id}/section"
        section_data = await make_get_request(section_url, token_manager, session)

        if not section_data:
            print(f"No sections found for requirement ID: {requirement_id}")
            continue

        for section in section_data:
            section_id = section.get("sectionId")
            section_description = section.get("description")
            policies_assigned_count = section.get("policiesAssignedCount", 0)

            summary_data.append([
                section_id,
                section_description,
                policies_assigned_count,
                requirement_name,
                requirement_description
            ])
    return summary_data


def generate_csv(customer_name, stnd_name, totals, maybe_totals, included_totals, custom_totals, alert_rule_totals, auto_dismiss_totals, alert_rules_dict):
    output = StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    # Write totals for policies and custom policies
    writer.writerow(["CloudType", "Included", "Policies", "Custom"])
    total_counts = defaultdict(int)
    for cloud_type in sorted(totals.keys()):
        included_data = totals[cloud_type]
        for included in sorted(included_data.keys()):
            count = included_data[included]
            custom_count = custom_totals[cloud_type].get(included, 0)
            # Replace `included` value with `stnd_name` if it matches `customer_name` for display purposes
            display_included = stnd_name if included == customer_name else included
            writer.writerow([cloud_type, display_included, count, custom_count])
            total_counts[included] += count
            total_counts[f'{included}_Custom'] += custom_count
    writer.writerow([])
    writer.writerow([f"{customer_name} Total Policies for all clouds"])
    for included, count in total_counts.items():
        if included.endswith("_Custom"):
            continue
        # Replace `included` value with `stnd_name` if it matches `customer_name` for display purposes
        display_included = stnd_name if included == customer_name else included
        writer.writerow(["All Clouds", f"Total {display_included}", count])
    total_policies = sum(total_counts[included] for included in total_counts if not included.endswith("_Custom"))
    writer.writerow(["All Clouds", "Total Policies", total_policies])
    print(f"{customer_name} Total Policies: {total_policies}")
    writer.writerow(["All Clouds", "Total Custom", sum(total_counts[included] for included in total_counts if included.endswith("_Custom"))])
    # Write totals for MAYBE and YES policies per cloud
    # writer.writerow([])
    # writer.writerow([f"Compliance Standard - {stnd_name}"])
    # writer.writerow(["CloudType", "Requirement/Section", "Value", "MAYBE Policies", "Included Policies"])
    for cloud_type in sorted(maybe_totals.keys() | included_totals.keys()):  # Ensure all cloud types are considered
        maybe_data = maybe_totals.get(cloud_type, {})
        included_data = included_totals.get(cloud_type, {})
        all_categories = set(maybe_data.keys()) | set(included_data.keys())  # Combine keys from both dictionaries
        for category in sorted(all_categories):  # Process each category
            maybe_values = maybe_data.get(category, {})
            included_values = included_data.get(category, {})
            all_values = set(maybe_values.keys()) | set(included_values.keys())  # Combine all values in the category
            for value in sorted(all_values):  # Process each value
                maybe_count = maybe_values.get(value, 0)  # Get count from maybe_totals
                included_count = included_values.get(value, 0)  # Get count from included_totals
                writer.writerow([cloud_type, category, value, maybe_count, included_count])
    writer.writerow([])

    # for cloud_type in sorted(maybe_totals.keys()):
    #     data = maybe_totals[cloud_type]
    #     for category, values in data.items():
    #         for value, maybe_count in values.items():
    #             included_count = included_totals[cloud_type][category].get(value, 0)
    #             writer.writerow([cloud_type, category, value, maybe_count])
    #     total_maybe_per_cloud = sum(maybe_totals[cloud_type]["Requirement"].values())
    #     writer.writerow([cloud_type, "Total", "", total_maybe_per_cloud])
    #     writer.writerow([])
    # writer.writerow(["CloudType", "Requirement/Section", "Value", "MAYBE Policies"])
    # for cloud_type in sorted(included_totals.keys()):
    #     data = included_totals[cloud_type]
    #     for category, values in data.items():
    #         for value, included_count in values.items():
    #             included_count = included_totals[cloud_type][category].get(value, 0)
    #             writer.writerow([cloud_type, category, value, included_count])
    #     total_included_per_cloud = sum(included_totals[cloud_type]["Requirement"].values())
    #     writer.writerow([cloud_type, "Total", "", total_included_per_cloud])
    #     writer.writerow([])
    # Write alert rule totals
    # Write detailed information for auto dismiss rules
    auto_dismiss_rules, total_unique_resource_list_auto_dismiss, total_auto_dismiss_rules = get_unique_resource_lists_count_for_auto_dismiss(alert_rules_dict)
    writer.writerow([])
    writer.writerow(["Alert Rules"])
    writer.writerow(["Alert Rule Name", "Policies"])
    for rule, count in alert_rule_totals.items():
        writer.writerow([rule, count])
    writer.writerow(["Auto Dismiss Rules:"])
    for rule in auto_dismiss_rules:
        writer.writerow([rule])
    writer.writerow(["Total Active Auto Dismiss Alert Rules", auto_dismiss_totals])
    writer.writerow(["Total Tag-based Resource Lists in Auto Dismiss Rules", total_unique_resource_list_auto_dismiss])
    return output.getvalue()


def process_consolidated_data(rows, headers, customer_name, standard_name, standard_summary_data):
    section_descriptions = {
        str(row[0]): row[1]
        for row in standard_summary_data[1:]
        if len(row) > 1 and row[0] and isinstance(row[0], str)
    }
    totals = defaultdict(lambda: defaultdict(int))
    maybe_totals = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    included_totals = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    custom_totals = defaultdict(lambda: defaultdict(int))
    alert_rule_totals = defaultdict(int)
    auto_dismiss_totals = 0
    cloud_type_index = headers.index("CloudType")
    included_index = headers.index("Included")
    requirement_index = headers.index("Requirement")
    section_index = headers.index("Section")
    ftypes_index = headers.index("FindingTypes")
    policy_mode_index = headers.index("PolicyMode")
    alert_rules_index = headers.index("Rules")
    standard_index = headers.index("Standard")

    for row in rows:
        cloud_type = row[cloud_type_index]
        included = row[included_index]
        if included == 'NO_SUPPORT':
            included = 'Standard Not Supported'
        # requirement = str(row[requirement_index]).strip()
        # p_section = str(row[section_index]).strip()
        # ftypes = row[ftypes_index]

        # print(f"Included: {included}, Requirement: {requirement}, CloudType: {cloud_type}")
        # section_description = section_descriptions.get(p_section, "Unknown Section")
        # p_section_with_description = f"{p_section} ({section_description})"
        policy_mode = row[policy_mode_index]
        if policy_mode != 'redlock_default':
            policy_mode = 'custom'
        alert_rules = row[alert_rules_index]

        if included in ["DUPLICATE", "DELETED"]:
            continue
        totals[cloud_type][included] += 1
        # # if standard and standard in standard_name:
        # # print(customer_name)
        # if included == "MAYBE" and (requirement or p_section):
        #     print(f"Included: {included}, Requirement: {requirement}, CloudType: {cloud_type}")
        #     maybe_totals[cloud_type]["Requirement"][requirement] += 1
        #     maybe_totals[cloud_type]["Section"][p_section_with_description] += 1
        # if included in [f'{customer_name}']:
        #     # print(f"Included: {included}, Requirement: {requirement}, CloudType: {cloud_type}")
        #     included_totals[cloud_type]["Requirement"][requirement] += 1
        #     included_totals[cloud_type]["Section"][p_section_with_description] += 1
        if "custom" in policy_mode:
            custom_totals[cloud_type][included] += 1
            totals[cloud_type][included] += 1
        # if not isinstance(ftypes_totals[cloud_type], dict):
        #     print(f"ERROR: ftypes_totals[{cloud_type}] is not a dict before processing ftypes.")
        #     raise ValueError(f"ftypes_totals[{cloud_type}] is not a dict: {ftypes_totals[cloud_type]}")
        # if ftypes:
        #     ftypes_list = [ftype.strip() for ftype in ftypes.split(",")]
        #     for ftype in ftypes_list:
        #         ftypes_totals[cloud_type][ftype] += 1
        if alert_rules:
            alert_rule_list = [rule.strip() for rule in alert_rules.split(",")]
            for rule in alert_rule_list:
                alert_rule_totals[rule] += 1
                if "(auto_dismiss)" in rule:
                    auto_dismiss_totals += 1
    return totals, maybe_totals, included_totals, custom_totals, alert_rule_totals, auto_dismiss_totals


# Asynchronous function to get policy RQL
async def get_policy_rql(baseurl, token_manager, search_id, session, retries=3):
    if not search_id:
        print(f"Invalid search_id provided: {search_id}")
        return None
    token = await token_manager.get_token()  # Correctly get the token
    search_url = f"{baseurl}/search/history/{search_id}"
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    for attempt in range(retries):
        try:
            async with session.get(search_url, headers=headers, ssl=SSL_VERIFY, timeout=60) as response:
                response_text = await response.text()
                # Handle a 404 status code by returning 'DELETED'
                if response.status == 404:
                    print(f"Policy not found (404) for search_id: {search_id} (attempt {attempt + 1})")
                    # return 'DELETED'
                    return 'Not Available'
                # Process successful response
                if response.status == 200:
                    policy_data = await response.json()
                    if policy_data.get('searchType') == 'asset':
                        policy_rql = policy_data.get('queryWithFindingNames', '')
                    elif policy_data.get('rule', {}).get('type') == 'drift':
                        policy_rql = "{\"category\":\"Drift\",\"resourceTypes\":[]}"
                    else:
                        policy_rql = policy_data.get('query', '')
                    return policy_rql
                else:
                    error_message = f"Error fetching policy RQL from API (attempt {attempt+1}): Status Code: {response.status})"
                    print(error_message)
                    print(f"Search URL: {search_url}, Search ID: {search_id}")
                    # print(f"Headers: {headers}")
                    print(f"Policy Response: {response_text}")
                # Retry after a delay if the status code is not 200
                await asyncio.sleep(2)
        except aiohttp.ClientError as e:
            print(f"ClientError on attempt {attempt+1}: {e}")
            await asyncio.sleep(2)
    # If all retries fail, return
    print(f"All {retries} attempts to fetch policy RQL failed for search_id: {search_id}.")
    return None


async def validate_custom_rql(baseurl, token_manager, policy_type, policy_rql, session):
    token = await token_manager.get_token()  # Correctly get the token
    validate_url = f"{baseurl}/policy/rule/validate"
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    payload = {
        "policyType": policy_type,
        "rule": {
            "criteria": policy_rql
        }
    }
    try:
        async with session.post(validate_url, headers=headers, json=payload, ssl=SSL_VERIFY) as response:
            response_text = await response.text()
            if response.status == 200 and response_text.strip() == '':
                # print("RQL validation successful.")
                return True
            else:
                raise Exception(f"Failed to validate RQL: {response_text}")
    except aiohttp.ClientError as e:  # Catch aiohttp exceptions
        print(Fore.RED + f"Error during API call: {e}.  {validate_url} ")
        print(Fore.WHITE)
        return None
    except Exception as e:  # General exception catch
        print(Fore.RED + f"RQL Validation Error: {e} {payload}")
        print(Fore.WHITE)
        return None

# This is used to update custom policies to include cloud type in name and description (if required)
# def prepend_custom_with_cloud_type(rows, headers):
#     cloud_type_index = headers.index("CloudType")
#     policy_upi_index = headers.index("Policy UPI")
#     policy_mode_index = headers.index("PolicyMode")
#     policy_name_index = headers.index("PolicyName")
#     description_index = headers.index("Description")
#     for row in rows:
#         policy_upi = row[policy_upi_index]
#         policy_mode = row[policy_mode_index]
#         # if policy_upi.startswith("Custom"):
#         if policy_mode == 'custom':
#             policy_name = row[policy_name_index]
#             description = row[description_index]
#             cloud_type = row[cloud_type_index].upper()
#             # print(f" Cloud - {cloud_type} Custom Policy name: {policy_name} - Description: {description}")
#             if not policy_name.lower().startswith(cloud_type.lower()):
#                 row[policy_name_index] = f"{cloud_type} {policy_name}"
#                 # print(f" New Custom Policy name: {row[policy_name_index]}")
#             if not description.lower().startswith(cloud_type.lower()):
#                 row[description_index] = f"{cloud_type} {description}"
#                 # print(f"New Custom Policy - Description:  {row[description_index]}")
#             if not policy_upi:
#                 row[policy_upi_index] = "Custom"


def update_labels(rows, headers, customer_name, stnd_name):
    updated_rows = []
    customer_ID = customer_name[:2].upper()  # Get the first two letters of Customer name in capitals
    included_index = headers.index("Included")
    labels_index = headers.index("Labels")
    # last_modified_by_index = headers.index("LastModifiedBy")
    cloud_type_index = headers.index("CloudType")
    p_section_index = headers.index("SectionOption")
    policy_mode_index = headers.index("PolicyMode")
    for row in rows:
        included = row[included_index]  # Using the Included column
        labels = row[labels_index]
        # last_modified_by = row[last_modified_by_index]
        cloud_type = row[cloud_type_index].upper()
        p_section = row[p_section_index]
        p_section = str(p_section)
        policy_mode = row[policy_mode_index]
        if included in ["MAYBE", "YES"]:  # Include both MAYBE and YES
            # Split labels and filter out any that start with {customer_ID}_
            labels_list = [label for label in labels.split(',') if label and not label.startswith(f'{customer_ID}_')]
            standard_label = stnd_name.replace(' ', '_')
            if standard_label not in labels_list:
                labels_list.append(standard_label)
            if p_section:
                # Ensure p_section is treated as a string, even if it's None or NaN
                p_section_label = str(p_section).replace(' ', '_') if p_section is not None else ''
                if p_section and f'{customer_ID}_{cloud_type}_{p_section_label}' not in labels_list:
                    labels_list.append(f'{customer_ID}_{cloud_type}_{p_section_label}')
            # Add Custom for rows that don't equal to LastModifiedBy = "Prisma Cloud System Admin" and don't already have Custom
            if policy_mode == "custom" and 'Custom' not in labels_list:
                labels_list.append('Custom')
            # Update the Labels column with the filtered and updated list
            row[labels_index] = ','.join(labels_list)
        updated_rows.append(row)
    return updated_rows


def extract_api_name(rql):
    if not rql:
        return ''
    patterns = {
        "from_api": r"(network|event)\s+from\s+([^\s]+)",
        "api_name": r"api\.name\s*=\s*'([^']*)'",
        "finding_source": r"finding\.source\s*=\s*'([^']*)'",
        "config_network": r"config\s+from\s+network",
        "finding_type": r"finding\.type\s*=\s*'([^']*)'",
        "config_iam": r"config\s+from\s+iam",
        "asset_type": r"asset\.type\s*(?:=\s*'([^']*)'|IN\s*\(\s*'([^']*)'\s*\))",
        "drift": r"Drift:([a-zA-Z]+)",
        "resource_types": r'"resourceTypes":\[(.*?)\]',
        "category": r'"category":"([^"]+)"',
        # "metadata": r"metadata:\s*.*?provider:\s*\"([^\"]*)\".*?category:\s*\"([^\"]*)\"",
        "metadata": r"metadata:\s*.*?provider:\s*([^\s]+).*?category:\s*([^\s]+)",
        # "custom_metadata": r"metadata:\s*.*?category:\s*\"([^\"]+)\".*?provider:\s*\"([^\"]+)\".*?resource_types:\s*-\s*\"([^\"]+)\"",
        "custom_metadata": r"metadata:\s*.*?category:\s*([^\s]+).*?provider:\s*([^\s]+).*?resource_types:\s*-\s*([^\s]+)",
        "terraform_module": r"resource_types:\s*-\s*\"([^\"]+)\".*?attribute:\s*([^\s]+)"
    }
    for key, pattern in patterns.items():
        match re.search(pattern, rql, re.DOTALL):
            case re.Match() as m if key == "from_api":
                return m.group(2)
            case re.Match() as m if key == "api_name":
                return m.group(1)
            case re.Match() as m if key == "finding_source":
                return m.group(1)
            case re.Match() as m if key == "config_network":
                return 'network_exposure'
            case re.Match() as m if key == "finding_type":
                return m.group(1)
            case re.Match() if key == "config_iam":
                return 'iam'
            case re.Match() as m if key == "asset_type":
                return m.group(1) or m.group(2)
            case re.Match() as m if key == "drift":
                return f"IaC:Drift:{m.group(1)}"
            case re.Match() as m if key == "resource_types":
                resource_types = m.group(1).replace('"', '').split(',')
                if resource_types and resource_types[0] == '*':
                    match re.search(patterns["category"], rql):
                        case re.Match() as cm:
                            return process_category(cm.group(1))
                else:
                    api_names = [
                        rt.strip().replace("_", "-").replace("azurerm", "azure")
                        for rt in resource_types
                    ]
                    api_name = ",".join(api_names)
                    match re.search(patterns["category"], rql):
                        case re.Match() as cm:
                            return f"{process_category(cm.group(1))}:{api_name}"
                    return api_name
            case re.Match() as m if key == "metadata":
                provider, category = m.groups()
                return f"{process_category(category)}:{provider}"
            case re.Match() as m if key == "custom_metadata":
                category, provider, resource_type = m.groups()
                return f"{process_category(category)}:{provider}:{resource_type}"
            case re.Match() as m if key == "terraform_module":
                resource_type, attribute = m.groups()
                return f"{resource_type}:{attribute}"
    # If no match is found, return an empty string
    return ''


def process_category(category):
    """
    Function to prepend appropriate prefix based on the category value.
    """
    code_categories = ["Sast", "Secrets", "Licenses"]
    cicd_categories = [
        "System Configuration", "Artifact Integrity Validation", "Credential Hygiene",
        "Data Protection", "Flow Control Mechanisms", "Input Validation",
        "Poisoned Pipeline Execution (PPE)", "Pipeline Flow Control",
        "Pipeline-Based Access Controls (PBAC)", "Supply Chain", "Dependency Chains"
    ]
    if category in code_categories:
        return f"Code:{category}"
    elif category in cicd_categories:
        return f"CI/CD:{category}"
    else:
        return f"IaC:{category}"


def derive_service_name(api):
    """
    Derives the service name from the API name. Handles cases where the input is None or invalid.
    """
    if not api:  # Handle None or empty API value
        return ''
    # Define regex patterns
    patterns = {
        "resource_types": r'"resourceTypes":\[(.*?)\]',
        "category": r'"category":"([^"]+)"'
    }
    # Match resourceTypes and category in the criteria
    match_resource_types = re.search(patterns["resource_types"], api)
    match_category = re.search(patterns["category"], api)
    # Use match-case for evaluating different conditions
    match match_resource_types, api:
        case re.Match() as rt_match, _:
            resource_types = rt_match.group(1).replace('"', '').split(',')
            if resource_types and resource_types[0]:  # If not empty
                service_names = set(derive_service_name(api) for api in resource_types)
                return ",".join(sorted(service_names))
        case None, _ if match_category and '[]' in api:
            return match_category.group(1)
        case None, _ if api.startswith("aws-"):
            parts = api.split("-")
            return "-".join(parts[:2])
        case None, _ if api.startswith("aws-storage-"):
            parts = api.split("-")
            return "-".join(parts[:3]) if len(parts) > 2 else api
        case None, _ if api.startswith("gcloud-cloud-"):
            parts = api.split("-")
            return f"gcloud-cloud-{parts[2]}" if len(parts) > 2 else api
        case None, _ if "api-key" in api:
            parts = api.split("-")
            api_key_index = parts.index('api') + 1
            return '-'.join(parts[:api_key_index + 1]) if len(parts) > api_key_index else api
        case None, _ if api.startswith("gcp-compute") or api.startswith("gcloud-compute"):
            parts = api.split('-')
            if api.startswith(("gcloud-compute-instance", "gcloud-compute-firewall", "gcloud-compute-networks")):
                return '-'.join(parts[:3]) if len(parts) > 2 else api
            return "gcloud-compute"
        case None, _ if any(term in api for term in ["app-engine", "cloud-run", "service-directory", "security-command-center", "block-storage", "object-storage", "app-stream", "file-storage"]):
            parts = api.split('-')
            return '-'.join(parts[:3]) if len(parts) > 2 else api
        case None, _ if api == "azure-kubernetes-cluster":
            return "azure-aks"
        case None, _ if api == "azure-container-instances-container-group":
            return "azure-aci"
        case None, _ if api.startswith("azure-"):
            parts = api.split('-')
            return '-'.join(parts[:3]) if len(parts) > 2 else api
        case None, _ if api.startswith("alibaba-cloud-"):
            parts = api.split("-")
            return "-".join(parts[:3]) if len(parts) > 2 else api
        case _:
            return '-'.join(api.split('-')[:2])


def get_unique_resource_lists_count_for_auto_dismiss(alert_rules_dict):
    unique_resource_lists = set()
    auto_dismiss_rules = []
    for rule_name, rule_details in alert_rules_dict.items():
        # Check if the rule name contains "(auto_dismiss)"
        if "(auto_dismiss)" in rule_name:
            # Extract resource lists from the rule details and add them to the set
            resource_lists = rule_details.get("resourceLists", [])
            unique_resource_lists.update(resource_lists)
            auto_dismiss_rules.append(rule_name)
    # Return the count of unique resource lists
    return auto_dismiss_rules, len(unique_resource_lists), len(auto_dismiss_rules)


def handle_build_children(children, is_custom=False):
    """
    Extract data from build children and return RQL, Checkov ID, and recommendation.
    For custom policies, if 'criteria' is missing, it returns data from 'metadata'.
    """
    for child in children:
        # Handle custom policies with 'metadata' instead of 'criteria'
        if is_custom and not child.get('criteria', ''):
            metadata = child.get('metadata', {})
            return metadata.get('code', ''), metadata.get('checkovId', ''), child.get('recommendation', '')
        # Handle standard policies with 'criteria'
        return child.get('criteria', ''), child.get('metadata', {}).get('checkovId', ''), child.get('recommendation', '')
    # Default return if no matching child found
    return '', '', ''


# Define a function to handle build and run policies
async def handle_build_and_run_policy(policy_data, baseurl, token_manager, session):
    search_id = policy_data.get('rule', {}).get('criteria', '')  # Run criteria
    api_name = policy_data.get('apiName')

    if not search_id and policy_data.get('rule', {}).get('type') == 'drift':
        # print(f"Problem search_id - {policy_data}")
        run_policy_rql = "{\"category\":\"Drift\",\"resourceTypes\":[]}"
    else:
        # Fetch run RQL using search_id
        run_policy_rql = await get_policy_rql(baseurl, token_manager, search_id, session)
    checkov_id, build_policy_rql, build_recommendation = None, '', ''
    recommendation = policy_data.get('recommendation', '')
    # Process build part
    children = policy_data.get('rule', {}).get('children', [])
    for child in children:
        if child.get('type') == 'build':
            # Extract build-related values
            build_policy_rql = child.get('criteria', '')
            checkov_id = child.get('metadata', {}).get('checkovId', '')
            build_recommendation = child.get('recommendation', '')
            break
    # Combine run and build RQL
    combined_policy_rql = f"{run_policy_rql}\r\n{build_policy_rql}".strip()
    # Extract API name and service name based on the combined RQL
    if not api_name:
        api_name = extract_api_name(build_policy_rql if not run_policy_rql else run_policy_rql)
    service_name = derive_service_name(api_name)
    # Prepare recommendation combining both run and build parts
    combined_recommendation = f"{recommendation} \r\nIaC: {build_recommendation}".strip()
    # Return the relevant fields for further processing or storing in the final output
    return run_policy_rql, combined_policy_rql, api_name, service_name, checkov_id, combined_recommendation


async def get_alerts_data(baseurl, token_manager, session, status=None):
    token = await token_manager.get_token()
    alerts_url = f"{baseurl}/alert/v1/policy"
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    # Create payload based on status
    payload = {
        "filters": [],
        "fields": ["policyId", "alertCount"],
        "timeRange": {
            "type": "to_now",
            "value": "epoch"
        }
    }
    if status:
        payload["filters"].append({
            "name": "alert.status",
            "value": status,
            "operator": "="
        })
    else:
        payload["filters"].append({
            "name": "alert.status",
            "value": "resolved",
            "operator": "="
        })
    async with session.post(alerts_url, headers=headers, json=payload, ssl=SSL_VERIFY) as response:
        if response.status != 200:
            error_message = f"Error fetching {status if status else 'resolved'} alert counts from API: {await response.text()} (Status Code: {response.status})"
            print(error_message)
            return []
        response_data = await response.json()
        return response_data.get("policies", [])


async def get_all_alerts_count(baseurl, token_manager, session):
    # Fetch all alert counts
    alert_count_data = await get_alerts_data(baseurl, token_manager,  session, status="resolved")
    alert_count_open_data = await get_alerts_data(baseurl, token_manager,  session, status="open")
    alert_count_dismissed_data = await get_alerts_data(baseurl, token_manager,  session, status="dismissed")
    # Create a dictionary with policyId as key and alerts_count, alerts_count_open, and alerts_count_dismissed as values
    alerts_count_dict = {}
    # Populate the dictionary with total alerts count
    for alert in alert_count_data:
        policy_id = alert.get("policyId")
        alerts_count_dict.setdefault(policy_id, {
            "alerts_count": 0,
            "alerts_count_open": 0,
            "alerts_count_dismissed": 0
        })
        alerts_count_dict[policy_id]["alerts_count"] = alert.get("alertCount", 0)
    # Update the dictionary with open alerts count
    for alert in alert_count_open_data:
        policy_id = alert.get("policyId")
        alerts_count_dict.setdefault(policy_id, {
            "alerts_count": 0,
            "alerts_count_open": 0,
            "alerts_count_dismissed": 0
        })
        alerts_count_dict[policy_id]["alerts_count_open"] = alert.get("alertCount", 0)
    # Update the dictionary with dismissed alerts count
    for alert in alert_count_dismissed_data:
        policy_id = alert.get("policyId")
        alerts_count_dict.setdefault(policy_id, {
            "alerts_count": 0,
            "alerts_count_open": 0,
            "alerts_count_dismissed": 0
        })
        alerts_count_dict[policy_id]["alerts_count_dismissed"] = alert.get("alertCount", 0)
    return alerts_count_dict


async def get_policies_alert_rules(baseurl, token_manager, session):
    alerts_rule_url = f"{baseurl}/v2/alert/rule"
    alerts_rule_data = await make_get_request(alerts_rule_url, token_manager, session)
    account_groups_url = f"{baseurl}/cloud/group"
    account_groups = await make_get_request(account_groups_url, token_manager, session)
    alerts_rules_dict = {}

    for rule in alerts_rule_data:
        rule_name = rule.get("name", "")
        scan_all = rule.get("scanAll", False)
        target_resource_list = rule.get("target", {}).get("targetResourceList", {})
        alert_rule_policy_filter = rule.get("alertRulePolicyFilter", {})

        # Append (auto_dismiss) only if action is AUTO_DISMISS, enabled is True, and ids list is not empty
        if (
            target_resource_list.get("action") == "AUTO_DISMISS" and
            target_resource_list.get("enabled") and
            isinstance(target_resource_list.get("ids"), list) and
            len(target_resource_list.get("ids")) > 0
        ):
            rule_name = f"{rule_name} (auto_dismiss)"
            print(f"adding auto_dismiss to {rule_name}")

        # Handle "scanAll" rules
        if scan_all:
            alerts_rules_dict[rule_name] = {
                "policies": "all",
                "notificationChannels": rule.get("notificationChannels", []),
                "accountGroupsName": [
                    group['name'] for group in account_groups
                    if group['id'] in rule.get("target", {}).get("accountGroups", [])
                ],
                "resourceLists": target_resource_list.get("ids", []),
                "alertRulePolicyFilter": alert_rule_policy_filter
            }
        else:
            policy_ids = rule.get("policies", [])
            # Create the alert rule entry in the dictionary for specific policies
            alerts_rules_dict[rule_name] = {
                "policies": policy_ids,
                "notificationChannels": rule.get("notificationChannels", []),
                "accountGroupsName": [
                    group['name'] for group in account_groups
                    if group['id'] in rule.get("target", {}).get("accountGroups", [])
                ],
                "resourceLists": target_resource_list.get("ids", []),
                "alertRulePolicyFilter": alert_rule_policy_filter
            }
    return alerts_rules_dict


async def get_alert_rule_names_for_policy(policy_id, cloud_type, severity, alerts_rules_dict):
    matching_rules = []
    for rule_name, rule_details in alerts_rules_dict.items():
        policies = rule_details.get("policies", [])
        alert_rule_filter = rule_details.get("alertRulePolicyFilter", {})
        # Check if policy id matches, rule is "scanAll", or filter matches severity and cloud type
        if (
            policies == "all" or
            policy_id in policies or
            (
                severity in alert_rule_filter.get("policy.severity", []) and
                cloud_type in alert_rule_filter.get("cloud.type", [])
            )
        ):
            matching_rules.append(rule_name)

            # Log matching rules with auto_dismiss for clarity
            if "auto_dismiss" in rule_name:
                print(f"Matching rule with auto_dismiss: {rule_name}")
    # Sort the matching rules by name before returning
    matching_rules.sort()
    return ", ".join(matching_rules)


async def get_AccountGroups_for_policy(policy_id, alerts_rules_dict):
    matching_rules = []
    for rule_name, rule_details in alerts_rules_dict.items():
        if rule_details["policies"] == "all" or policy_id in rule_details["policies"]:
            notification_channels = rule_details.get("notificationChannels", [])
            account_groups = rule_details.get("accountGroupsName", [])
            details = f"{rule_name} - Notifications:{','.join(notification_channels)}, AccountGroups:{','.join(account_groups)};"
            matching_rules.append(details)
    matching_rules.sort()
    return ", ".join(matching_rules)



async def update_policies(baseurl, token_manager, rows, headers, alerts_count_dict, alerts_rules_dict, cloud_types, session):
    """
    Update Policy RQL, API Name, Service, and Policy UPI values for non-build policies.
    """
    # Extract index positions from headers
    policy_id_index = headers.index("PolicyID")
    policy_rql_index = headers.index("Policy RQL")
    api_name_index = headers.index("API Name")
    # policy_type_index = headers.index("PolicyType")
    policy_name_index = headers.index("PolicyName")
    policy_description_index = headers.index("Description")
    policy_severity_index = headers.index("Severity")
    policy_subtypes_index = headers.index("PolicySubTypes")
    # requirement_index = headers.index("Requirement")
    policy_upi_index = headers.index("Policy UPI")
    included_index = headers.index("Included")
    search_id_index = headers.index("search_id")
    service_index = headers.index("Service")
    enabled_index = headers.index("Enabled")
    last_modified_by_index = headers.index("LastModifiedBy")
    recommendation_index = headers.index("Recommendation")
    remediation_index = headers.index("Remediation")
    remediable_index = headers.index("Remediable")
    last_modified_on_index = headers.index("LastModifiedOn")
    policyCategory_index = headers.index("PolicyCategory")
    findingTypes_index = headers.index("FindingTypes")
    policyClass_index = headers.index("PolicyClass")
    owner_index = headers.index("Owner")
    reason_index = headers.index("Reason")
    system_default_index = headers.index("SystemDefault")
    cloud_type_index = headers.index("CloudType")
    labels_index = headers.index("Labels")
    policy_mode_index = headers.index("PolicyMode")
    alerts_count_index = headers.index("Resolved")
    alerts_count_open_index = headers.index("Open")
    alerts_count_dismissed_index = headers.index("Dismissed")
    alert_rules_index = headers.index("Rules")
    AccountGroups_index = headers.index("AccountGroups")
    # Track existing policies
    existing_policy_ids = {row[policy_id_index] for row in rows}

    async def update_rows_with_policy_data(row, policy_data, alerts_count_dict, alerts_rules_dict, session):
        """
        Helper function to update a row based on the policy data.
        """
        # Extract main fields from the policy data
        system_default = policy_data.get('systemDefault', '')
        cloud_type = policy_data.get('cloudType', '')
        labels = ",".join(policy_data.get('labels', [])) if policy_data.get('labels') else ''
        policy_mode = policy_data.get('policyMode', '')
        policy_id = policy_data.get('policyId', '')
        policy_type = policy_data.get('policyType', '')
        # policy_subtype = policy_data.get('policySubTypes', [''])[0]
        policy_subtypes = policy_data.get('policySubTypes', [])
        policy_subtypes_str = ",".join(policy_subtypes) if policy_subtypes else ''
        policy_name = policy_data.get('name', '')
        description = policy_data.get('description', '')
        severity = policy_data.get('severity', '')
        last_modified_by = policy_data.get('lastModifiedBy', '')
        recommendation = policy_data.get('recommendation', '')
        remediation_value = policy_data.get('remediation', '')
        remediation = str(remediation_value).replace('{', '').replace('}', '').replace(', ', ',\r\n')
        remediable = policy_data.get('remediable', '')
        last_modified_on = policy_data.get('lastModifiedOn', 0)
        enabled = policy_data.get('enabled', False)
        policy_upi = policy_data.get('policyUpi', '')
        api_name = policy_data.get('apiName')
        alerts_count = alerts_count_open = alerts_count_dismissed = ''
        if enabled:
            if alerts_data := alerts_count_dict.get(policy_id):
                alerts_count = alerts_data.get("alerts_count", 0)
                alerts_count_open = alerts_data.get("alerts_count_open", 0)
                alerts_count_dismissed = alerts_data.get("alerts_count_dismissed", 0)
        alert_rules = await get_alert_rule_names_for_policy(policy_id, cloud_type, severity, alerts_rules_dict)
        if include_AccountGroups:
            AccountGroups = await get_AccountGroups_for_policy(policy_id, alerts_rules_dict)
        else:
            AccountGroups = ''
            # Convert timestamp to a readable datetime format
        last_modified_on_str = datetime.fromtimestamp(last_modified_on / 1000).strftime('%Y-%m-%d %H:%M:%S') if last_modified_on else ''
        policyCategory = policy_data.get('policyCategory', '')
        findingTypes = ",".join(str(findingType) for findingType in policy_data.get('findingTypes', []) if findingType) if policy_data.get('findingTypes') else ''
        policyClass = policy_data.get('policyClass', '')
        owner = policy_data.get('owner', '')
        search_id = policy_data.get('rule', {}).get('criteria', '')
        checkov_id, build_policy_rql, build_recommendation, api_name, service_name,  = None, '', '', '', ''
        # Map certain policy types to specific API names and service names
        type_service_map = {
            'iam': (f"{policy_type}", f"{policy_subtypes_str}_{policy_type}"),
            'audit_event': (f"{policy_type}", f"{policy_type}"),
            'anomaly': (f"{policy_type}_{policy_subtypes_str}", f"{search_id}"),
            'api': ('network_event_api', f"{policy_subtypes_str}_{policy_type}"),
            'malware': (f"{policy_type}", f"{policy_type}_{policy_subtypes_str}"),
            'grayware': (f"{policy_type}", f"{policy_type}_{policy_subtypes_str}"),
            'data': (f"{policy_type}", f"{policy_subtypes_str}_{policy_type}"),
            'workload_vulnerability': (f"{policy_type}", f"{policy_type}"),
            'workload_incident': (f"{policy_type}", f"{policy_type}")
        }
        # Process common RQL handling for certain policy subtypes
        policy_subtypes_to_rql = ['run', 'audit', 'network_event', 'event', 'network_exposure', 'network_config', 'misconfig', 'misconfig_and_event', 'permissions']
        # Handle build policies
        if policy_subtypes_str == 'build':
            children = policy_data.get('rule', {}).get('children', [])
            policy_rql, checkov_id, build_recommendation = handle_build_children(
                children, is_custom=(policy_data.get('policyMode') == 'custom'))
            recommendation = build_recommendation
            if checkov_id:
                policy_upi = checkov_id
            if not api_name:
                api_name = extract_api_name(policy_rql)
            service_name = derive_service_name(api_name)
        # Handle run_and_build policies
        elif policy_subtypes_str == 'build,run':
            run_policy_rql, policy_rql, api_name, service_name, checkov_id, recommendation = await handle_build_and_run_policy(policy_data, baseurl, token_manager, session)
            # Update Policy UPI with Checkov ID if applicable
            if checkov_id and checkov_id not in policy_upi:
                policy_upi = f"{policy_upi}-{checkov_id}" if policy_upi else checkov_id
        elif search_id and policy_subtypes_str in policy_subtypes_to_rql:
            # Fetch the RQL for the specific subtype
            policy_rql = await get_policy_rql(baseurl, token_manager, search_id, session)
            # Extract API name and service name if RQL exists
            if not api_name:
                api_name = extract_api_name(policy_rql)
            service_name = derive_service_name(api_name)
        else:
            # Fallback to use the search_id if no specific RQL is found
            policy_rql = search_id
            # Ensure the default return is a tuple
            if not api_name:
                api_name, service_name = type_service_map.get(policy_type, (api_name, service_name))
        # Handle specific policies that don't support standards
        if policy_type in ['attack_path', 'iam', 'api', 'workload_vulnerability', 'workload_incident'] or policy_subtypes_str in ["network_config"]:
            row[included_index] = "NO_SUPPORT"
        timestamp = datetime.now().strftime('%Y/%m/%d/%H:%M')
        if policy_rql not in ['', 'Not Available'] and policy_mode == 'custom' and policy_subtypes_str != 'build':
            if policy_subtypes_str == 'build,run' and run_policy_rql:
                rql_valid = await validate_custom_rql(baseurl, token_manager, policy_type, run_policy_rql, session)
            else:
                rql_valid = await validate_custom_rql(baseurl, token_manager, policy_type, policy_rql, session)
            reason = f"Policy details Updated: {rql_valid}" + timestamp
            if rql_valid == "RQL validation successful" and policy_upi == '':
                policy_upi = 'Custom'
            else:
                policy_upi = "Custom-Review"
        else:
            reason = "Policy details Updated" + timestamp
        # Update the row values
        row[api_name_index] = api_name
        row[service_index] = service_name
        row[search_id_index] = search_id
        row[policy_rql_index] = policy_rql
        row[policy_upi_index] = policy_upi
        # row[requirement_index] = requirement
        row[policy_subtypes_index] = policy_subtypes_str
        row[policy_name_index] = policy_name
        row[policy_description_index] = description
        row[policy_severity_index] = severity
        row[enabled_index] = enabled
        row[last_modified_by_index] = last_modified_by
        row[recommendation_index] = recommendation
        row[remediation_index] = remediation
        row[remediable_index] = remediable
        row[last_modified_on_index] = last_modified_on_str
        row[policyCategory_index] = policyCategory
        row[findingTypes_index] = findingTypes
        row[policyClass_index] = policyClass
        row[owner_index] = owner
        row[reason_index] = reason
        row[system_default_index] = system_default
        row[cloud_type_index] = cloud_type
        row[labels_index] = labels
        row[policy_mode_index] = policy_mode
        row[alerts_count_index] = alerts_count
        row[alerts_count_open_index] = alerts_count_open
        row[alerts_count_dismissed_index] = alerts_count_dismissed
        row[alert_rules_index] = alert_rules
        row[AccountGroups_index] = AccountGroups


    async def update_rows_with_with_deleted(row, policy_id, status, reason):
        """
        Helper function to update a row with policy that no longer exists.
        """
        # Update the row values
        row[included_index] = status
        row[policy_id_index] = policy_id
        row[reason_index] = reason
    # Check and fetch policies for relevant subtypes
    policy_subtypes_to_check = ['run', 'run_and_build', 'build', 'audit', 'network_event', 'network', 'network_config',
                                'ueba', 'dns', 'event', 'identity', 'permissions', 'misconfig', 'misconfig_and_event']
    all_existing_policies = []
    # Iterate over each policy type to fetch relevant policies and filter based on existing policy IDs
    for policy_subtype in policy_subtypes_to_check:
        policy_url = f"{baseurl}/v2/policy?cloud.type={cloud_types}&policy.subtype={policy_subtype}"
        policies = await make_get_request(policy_url, token_manager, session)
        filtered_policies = [policy for policy in policies if policy["policyId"] in existing_policy_ids]  # Only keep policies that already exist
        all_existing_policies.extend(filtered_policies)
        print(f"Existing {policy_subtype} Policies to update: {len(filtered_policies)}")
    print(f"Existing Total Policies to update for cloud types: {cloud_types}: {len(all_existing_policies)}")
    # Create a dictionary for quick lookup by policyId
    policy_lookup = {policy["policyId"]: policy for policy in all_existing_policies}
    # Update existing rows based on the filtered policy data
    for row in tqdm(rows, desc="Updating Existing Rows"):
        policy_id = row[policy_id_index]
        policy_upi = row[policy_upi_index]
        # Ensure policy_id is non-empty before processing
        if policy_id:
            policy_data = policy_lookup.get(policy_id)  # Fetch policy data from the lookup dictionary
            if policy_data:
                # If policy data is found for the existing row, update it
                await update_rows_with_policy_data(row, policy_data, alerts_count_dict, alerts_rules_dict, session)
            else:
                # If the policy data does not exist in the lookup, mark it as deleted
                reason = f"Policy modified/deleted {datetime.now().strftime('%Y/%m/%d/%H:%M')}"
                status = "DELETED"
                await update_rows_with_with_deleted(row, policy_id, status, reason)


async def add_new_policy_to_rows(policy_data, baseurl, token_manager, rows, headers, alerts_count_dict, alerts_rules_dict, session):
    """
    Adds a new policy to the rows list based on the given policy data.
    """
    policy_id = policy_data.get('policyId', '')
    included_index = headers.index("Included")
    reason_index = headers.index("Reason")
    policy_name_index = headers.index("PolicyName")
    policy_mode_index = headers.index("PolicyMode")
    new_policy_name = policy_data.get('policyName', '')
    # Default values for included and reason
    included = f"{policy_data.get('policySubTypes', [''])[0]}-new".upper()
    reason = f"New Policy Found {datetime.now().strftime('%Y/%m/%d/%H:%M')}"
    # Check if there's an existing row with the same policy name and policy_mode == 'custom'
    for row in rows:
        # print(f"Custom policy: {policy_id} - {row[policy_mode_index]}")

        if row[policy_name_index] == new_policy_name and row[policy_mode_index] == 'custom':
            # If found, retain the existing values for `Included` and `Reason`
            included = row[included_index]
            reason = row[reason_index]
            print(f"Custom policy {included} : {reason}")

            break
    # Continue with the rest of the function as before, using the updated `included` and `reason`
    # Extract main fields from the policy data
    system_default = policy_data.get('systemDefault', '')
    cloud_type = policy_data.get('cloudType', '')
    labels = ",".join(policy_data.get('labels', [])) if policy_data.get('labels') else ''
    policy_mode = policy_data.get('policyMode', '')
    policy_type = policy_data.get('policyType', '')
    # policy_subtype = policy_data.get('policySubTypes', [''])[0]
    policy_subtypes = policy_data.get('policySubTypes', [])
    policy_subtypes_str = ",".join(policy_subtypes) if policy_subtypes else ''
    policy_name = policy_data.get('name', '')
    description = policy_data.get('description', '')
    severity = policy_data.get('severity', '')
    policy_upi = policy_data.get('policyUpi', '')
    last_modified_by = policy_data.get('lastModifiedBy', '')
    recommendation = policy_data.get('recommendation', '')
    remediation_value = policy_data.get('remediation', '')
    remediation = str(remediation_value).replace('{', '').replace('}', '').replace(', ', ',\r\n')
    remediable = policy_data.get('remediable', '')
    last_modified_on = policy_data.get('lastModifiedOn', 0)
    enabled = policy_data.get('enabled', False)
    api_name = policy_data.get('apiName')
    alerts_count = alerts_count_open = alerts_count_dismissed = ''
    if enabled:
        if alerts_data := alerts_count_dict.get(policy_id):
            alerts_count = alerts_data.get("alerts_count", 0)
            alerts_count_open = alerts_data.get("alerts_count_open", 0)
            alerts_count_dismissed = alerts_data.get("alerts_count_dismissed", 0)
    alert_rules = await get_alert_rule_names_for_policy(policy_id, cloud_type, severity, alerts_rules_dict)
    if include_AccountGroups:
        AccountGroups = await get_AccountGroups_for_policy(policy_id, alerts_rules_dict)
    else:
        AccountGroups = ''
        # Convert timestamp to a readable datetime format
    last_modified_on_str = datetime.fromtimestamp(last_modified_on / 1000).strftime('%Y-%m-%d %H:%M:%S') if last_modified_on else ''
    policyCategory = policy_data.get('policyCategory', '')
    findingTypes = ",".join(str(findingType) for findingType in policy_data.get('findingTypes', []) if findingType) if policy_data.get('findingTypes') else ''
    policyClass = policy_data.get('policyClass', '')
    owner = policy_data.get('owner', '')
    search_id = policy_data.get('rule', {}).get('criteria', '')
    checkov_id, build_policy_rql, build_recommendation, service_name = None, '', '', ''
    # Map certain policy types to specific API names and service names
    type_service_map = {
        'iam': (f"{policy_type}", f"{policy_subtypes_str}_{policy_type}"),
        'audit_event': (f"{policy_type}", f"{policy_type}"),
        'anomaly': (f"{policy_type}_{policy_subtypes_str}", f"{search_id}"),
        'api': ('network_event_api', f"{policy_subtypes_str}_{policy_type}"),
        'malware': (f"{policy_type}", f"{policy_type}_{policy_subtypes_str}"),
        'grayware': (f"{policy_type}", f"{policy_type}_{policy_subtypes_str}"),
        'data': (f"{policy_type}", f"{policy_subtypes_str}_{policy_type}"),
        'workload_vulnerability': (f"{policy_type}", f"{policy_type}"),
        'workload_incident': (f"{policy_type}", f"{policy_type}")
    }
    # Process common RQL handling for certain policy subtypes
    policy_subtypes_to_rql = ['run', 'audit', 'network_event', 'event', 'network_exposure', 'network_config', 'misconfig', 'misconfig_and_event', 'permissions']
    # Handle build policies
    if policy_subtypes_str == 'build':
        children = policy_data.get('rule', {}).get('children', [])
        policy_rql, checkov_id, build_recommendation = handle_build_children(
            children, is_custom=(policy_data.get('policyMode') == 'custom'))
        recommendation = build_recommendation
        if checkov_id:
            policy_upi = checkov_id
        if not api_name:
            api_name = extract_api_name(policy_rql)
        service_name = derive_service_name(api_name)
    elif policy_subtypes_str == 'build,run':
        run_policy_rql, policy_rql, api_name, service_name, checkov_id, recommendation = await handle_build_and_run_policy(policy_data, baseurl, token_manager, session)
        if checkov_id and checkov_id not in policy_upi:
            policy_upi = f"{policy_upi}-{checkov_id}" if policy_upi else checkov_id
    elif search_id and policy_subtypes_str in policy_subtypes_to_rql:
        policy_rql = await get_policy_rql(baseurl, token_manager, search_id, session)
        if not api_name:
            api_name = extract_api_name(policy_rql)
        service_name = derive_service_name(api_name)
    else:
        policy_rql = search_id
        if not api_name:
            api_name, service_name = type_service_map.get(policy_type, (api_name, service_name))
    if policy_type in ['attack_path', 'iam', 'api', 'workload_vulnerability', 'workload_incident']:
        included = "NO_SUPPORT-NEW"
    # Prepare the new row data
    new_row = [
        "", included, "NO", alerts_count, alerts_count_open, alerts_count_dismissed, enabled, alert_rules, cloud_type, service_name, policy_name, findingTypes, severity, owner,
        policy_type, policy_subtypes_str, "", "", labels, remediable, policyClass,
        api_name, description, recommendation, remediation, policy_rql, last_modified_by, last_modified_on_str, policy_upi,
        system_default, policy_mode, policyCategory, AccountGroups, policy_id, search_id, "true", reason
    ]
    # Add the new row to the rows list
    rows.append(new_row)
    # print(f">>Added new {included} policy: {policy_name}")


async def create_custom_search(baseurl, token_manager, policy_name, policy_rql, cloud_type, policy_type, session):
    token = await token_manager.get_token()
    if policy_type in ['network']:
        search_url = f"{baseurl}/search"
    elif policy_type in ['audit_event']:
        search_url = f"{baseurl}/search/event"
    else:
        # search_url = f"{baseurl}/search/config"
        search_url = f"{baseurl}/search/api/v2/config"
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    payload = {
        "limit": 10,
        "query": policy_rql,
        "heuristicSearch": "true",
        "timeRange": {
            "relativeTimeType": "BACKWARD",
            "type": "relative",
            "value": {
                "amount": 1,
                "unit": "hour"
            }
        }
    }
    print(f">>Creating search for policy: {policy_name}")
    # print(f'Payload for custom filter (before explicit JSON):  - {payload}')
    json_payload = json.dumps(payload, ensure_ascii=False).encode('utf-8')
    # print(f'{search_url} JSON Payload (explicitly encoded): - {json_payload}')
    try:
        async with session.post(search_url, headers=headers, data=json_payload, ssl=SSL_VERIFY) as response:
            response_text = await response.text()
            # if response.status == 200 and response_text.strip() == '':
            if response.status == 200:
                response_json = await response.json()
                search_id = response_json.get('id')
                print(f">>Search created successfully. Search ID: {search_id}")
                return search_id
            else:
                raise Exception(f"Failed to create search: {response_text}")
                return None
    except aiohttp.ClientError as e:  # Catch aiohttp exceptions
        print(Fore.RED + f">>Error during API call: {e}.  {search_url}")
        print(Fore.WHITE)
        return None
    except Exception as e:  # General exception catch
        print(Fore.RED + f">>Unexpected error: {e}.  {search_url}")
        print(Fore.WHITE)
        return None


async def create_custom_policy(baseurl, token_manager, policy_data, search_id, compliance_metadata, labels, session):
    token = await token_manager.get_token()
    policy_url = f"{baseurl}/policy"
    policy_type = policy_data['PolicyType']
    if policy_type in ['audit_event']:
        policy_rule_type = "AuditEvent"
    else:
        policy_rule_type = policy_type.capitalize()  # Capitalize policy type for rule
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    # Prepare rule based on policy type
    rule = {
        "criteria": search_id,
        "name": policy_data['PolicyName'],
        "parameters": {"savedSearch": "true"}
    }
    if policy_data['PolicyType'].lower() != 'network':
        # Add type for all non-network policy types
        rule["type"] = policy_rule_type
    # Prepare policySubTypes
    policy_sub_types = []
    if policy_data['PolicyType'].lower() == 'config':
        policy_sub_types = ["run"]
    elif policy_data['PolicyType'].lower() == 'network':
        policy_sub_types = ["network_event"]
    payload = {
        "cloudType": policy_data['CloudType'],
        "name": policy_data['PolicyName'],
        "policyType": policy_type,
        "policySubTypes": policy_sub_types,  # Add policy subtypes
        "rule": rule,
        "severity": policy_data['Severity'],
        "complianceMetadata": compliance_metadata,
        "labels": labels,
        "enabled": True  # Set policy to enabled
    }
    # print(f"Creating policy: {policy_data['PolicyName']} with payload: {json.dumps(payload, indent=2)}")
    print(f">>Creating policy: {policy_data['PolicyName']} ")
    try:
        async with session.post(policy_url, headers=headers, json=payload, ssl=SSL_VERIFY) as response:
            response_text = await response.text()
            if response.status == 200:
                response_json = await response.json()
                policy_id = response_json.get('policyId')
                print(f">>Policy created successfully. Policy ID: {policy_id}")
                return policy_id, response_text
    except aiohttp.ClientError as e:  # Catch aiohttp exceptions
        print(Fore.RED + f">>Error during API call: {e}.  {policy_url} ")
        print(Fore.WHITE)
        return None, e
    except Exception as e:  # General exception catch
        print(Fore.RED + f">>Unexpected error: {e}.  ")
        print(Fore.WHITE)
        return None, e


def check_existing_search(search_data_dict, policy_name):
    # Check in the pre-fetched dictionary for existing saved searches
    return search_data_dict.get(policy_name, '')


def filter_compliance_metadata(compliance_metadata, standard_id, section_id):
    # print(Fore.WHITE + f"Filtering for: {standard_id} - Section: {section_id} ")
    filtered_metadata = []
    for standard, metadata_list in compliance_metadata.items():
        for metadata in metadata_list:
            if metadata.get('standardId') == standard_id and metadata.get('sectionId') == section_id:
                filtered_metadata.append(metadata)
    return filtered_metadata


async def process_custom_policy(baseurl, token_manager, policy_data, sheet_headers, customer_name, search_data_dict, standard_id, stnd_name, compliance_metadata, session):
    timestamp = datetime.now().strftime('%Y/%m/%d/%H:%M')
    customer_ID = customer_name[:2].upper()  # Get the first two letters of Customer name in capitals
    policy_name = policy_data[sheet_headers.index('PolicyName')]
    policy_id = policy_data[sheet_headers.index('PolicyID')]
    policy_rql = policy_data[sheet_headers.index('Policy RQL')]
    cloud_type = policy_data[sheet_headers.index('CloudType')]
    p_section = policy_data[sheet_headers.index('SectionOption')]  # Extract Section ID
    policy_type = policy_data[sheet_headers.index('PolicyType')]
    policy_subtypes = policy_data[sheet_headers.index('PolicySubTypes')]
    if not p_section:
        reason = f"Failed to create - Missing Standard Section " + timestamp
        print(Fore.RED +f">>Failed to create policy - Missing Standard Section - 'PolicyName': {policy_name} - PolicyType': {policy_type},'PolicySubTypes': {policy_subtypes}, ")
        print(Fore.WHITE)
        return {'PolicyType': policy_type, 'PolicyName': policy_name, 'Reason': reason}
    # elif policy_subtypes in 'build':
    #     reason = f"Failed to create policy - Build policy is not supported {datetime.now().strftime('%Y%m%d%H%M')}"
    #     print(Fore.RED + f"Failed to create policy - Build policy is not supported - 'PolicyName': {policy_name} - PolicyType': {policy_type},'PolicySubTypes': {policy_subtypes}, ")
    #     print(Fore.WHITE)
    #     return {'PolicyType': policy_type, 'PolicyName': policy_name, 'Reason': reason}
    else:
        severity = policy_data[sheet_headers.index('Severity')]
        # Set enabled to True for new policy creation
        enabled = True
        labels = policy_data[sheet_headers.index('Labels')]
        # Process Labels
        # Ensure `labels` içs always a list
        if isinstance(labels, str):
            labels_list = [label.strip() for label in labels.split(',') if label.strip()]
        elif isinstance(labels, list):
            labels_list = labels
        else:
            labels_list = []
        # Filter out labels with `customer_ID` prefix
        labels_list = [label for label in labels_list if label and not label.startswith(f'{customer_ID}_')]
        # Process standard and section labels
        standard_label = stnd_name.replace(' ', '_')
        # Ensure p_section is treated as a string, even if it's None or NaN
        p_section_label = str(p_section).replace(' ', '_') if p_section is not None else ''
        if standard_label not in labels_list:
            labels_list.append(standard_label)
        if p_section and f'{customer_ID}_{cloud_type.upper()}_{p_section_label}' not in labels_list:
            labels_list.append(f'{customer_ID}_{cloud_type.upper()}_{p_section_label}')
        if 'Custom' not in labels_list:
            labels_list.append('Custom')
        policy_name_url = f"{baseurl}/v2/policy?policy.name={policy_name}"
        existing_policy_data = await make_get_request(policy_name_url, token_manager, session)
        if existing_policy_data:
            policy_data = existing_policy_data[0]
            # print(policy_data)
            existing_policy_id = policy_data.get('policyId')
            existing_search_id = policy_data.get('rule', {}).get('criteria')
            # Option to update other fields for existing custom policy

            print(f">>Updating spreadsheet with new details - existing custom Policy {policy_name}")
            reason = f"Updated " + timestamp
            return {'PolicyName': policy_name, 'PolicyID': existing_policy_id, 'search_id': existing_search_id, 'savedSearch': 'TRUE', 'Labels': labels_list, 'Reason': reason}
        # Check if saved search already exists
        existing_search_id = check_existing_search(search_data_dict, policy_name)
        if existing_search_id:
            print(f"Search for {policy_name} already exists with ID: {existing_search_id}")
            search_id = existing_search_id
        else:
            # Validate RQL
            try:
                await validate_custom_rql(baseurl, token_manager, policy_type, policy_rql, session)
            except Exception as e:
                print(f">>Failed to validate RQL for {policy_name}. Error: {str(e)}")
                reason = f"Failed to validate RQL " + timestamp

                return {'PolicyName': policy_name, 'search_id': '', 'savedSearch': '', 'PolicyID': '', 'Reason': reason}

            # Create a search if validation is successful
            try:
                search_id = await create_custom_search(baseurl, token_manager, policy_name, policy_rql, cloud_type, policy_type, session)
                # reason = f"Search created successfully {datetime.now().strftime('%Y%m%d%H%M')}"
                #
                # return {'PolicyType': policy_type, 'PolicyName': policy_name, 'search_id': search_id, 'savedSearch': 'TRUE', 'PolicyID': '', 'Reason': reason}

            except Exception as e:
                print(f">>Failed to create search for {policy_type} - {policy_name}. Error: {str(e)}")
                reason = f"Failed to create search - Error: {str(e)} " + timestamp
                return {'PolicyType': policy_type, 'PolicyName': policy_name, 'search_id': '', 'savedSearch': '', 'PolicyID': '', 'Reason': reason}
        # # Process labels
        # labels_list = [label.strip() for label in labels.split(',') if label.strip()]
        # Get filtered compliance metadata
        new_compliance_metadata = filter_compliance_metadata(compliance_metadata, standard_id, p_section)
        # Append the policyId to the new compliance metadata
        for metadata in new_compliance_metadata:
            metadata['policyId'] = ''  # Reset since policy_id is not known yet
            metadata.pop('sectionViewOrder', None)
            metadata.pop('requirementViewOrder', None)
            metadata.pop('systemDefault', None)
            metadata['customAssigned'] = "true"
        # print(f"New compliance: {new_compliance_metadata}")
        # Create a new policy

        try:
            policy_id, response_text = await create_custom_policy(baseurl, token_manager, {
                'CloudType': cloud_type,
                'PolicyName': policy_name,
                'PolicyType': policy_type,
                'Severity': severity,
                'Enabled': enabled,  # Set policy to enabled
                'policyUpi': 'Custom'

            }, search_id, new_compliance_metadata, labels_list, session)
            # success = f"Policy {policy_name} created successfully with Policy ID: {policy_id}"
            # print(success)
            if policy_id:
                reason = f"Created successfully " + timestamp
                return {'PolicyName': policy_name, 'search_id': search_id, 'savedSearch': 'TRUE', 'PolicyID': policy_id, 'Labels': labels_list, 'Reason': reason}
            else:
                reason = f"Failed to create - {response_text} " + timestamp
                return {'PolicyName': policy_name, 'search_id': search_id, 'savedSearch': 'TRUE', 'PolicyID': policy_id, 'Reason': reason}
        except Exception as e:
            print(f">>Failed to create policy for {policy_type} - {policy_name}. Error: {str(e)}")
            return {'PolicyType': policy_type, 'PolicyName': policy_name, 'Reason': f'Failed to create: {str(e)}'}


async def get_requirements(baseurl, token_manager, standard_id, p_sections, p_section, session):
    # Dictionary to store mapping of p_section to (requirementName, compliance_id)
    requirements_dict = {}
    matched_sections = set()  # Keep track of section_ids that found a match

    # Fetch all requirements for the given standard_id
    requirement_url = f"{baseurl}/compliance/{standard_id}/requirement"
    requirements_data = await make_get_request(requirement_url, token_manager, session)
    # print(f"{standard_id} {p_sections} Requirement Data: {requirements_data}")
    # Extract requirement IDs from the data
    requirement_ids = [req.get('id') for req in requirements_data if req.get('id')]
    # print(f"Requirement IDs: {requirement_ids}")
    # Convert p_sections to strings and strip whitespace for consistent comparison
    p_sections = [str(p).strip() for p in p_sections]

    # Iterate over each requirement_id
    for requirement_id in requirement_ids:
        # Fetch sections for the current requirement
        section_url = f"{baseurl}/compliance/{requirement_id}/section"
        section_data = await make_get_request(section_url, token_manager, session)
        # print(f"Section data for requirement ID {requirement_id}: {section_data}")
        # Process section_data as a list of dictionaries
        for section in section_data:
            section_id = str(section.get('sectionId', '')).strip()  # Convert sectionId to string and strip
            requirement_name = section.get('requirementName', '').strip()
            compliance_id = section.get('id')
            # Debug each potential match
            # print(f"Checking if sectionId '{section_id}' matches any p_section: {p_sections}")
            # Match sectionId with p_sections
            if section_id in p_sections:  # Both are strings, comparison will work
                # print(f"Matched sectionId: {section_id} with p_sections: {p_sections}")
                requirements_dict[section_id] = (requirement_name, compliance_id)
                matched_sections.add(section_id)

    # Check for any p_section that did not find a match
    unmatched_p_sections = set(p_sections) - matched_sections
    if unmatched_p_sections and p_section != 'section_id':
        for unmatched_section in unmatched_p_sections:
            print(f"Section ID '{unmatched_section}' from your input does not exist in the Compliance standard.")
        exit(1)

    # Debug output for the final dictionary
    # for p_section, (requirement_name, compliance_id) in requirements_dict.items():
    #     print(f"Matched Section ID: {p_section}, Requirement: {requirement_name}, Compliance ID: {compliance_id}")
    return requirements_dict


async def process_updated_policy(baseurl, token_manager, policy_data, sheet_headers, customer_name, included_value, standard_id, stnd_name, requirements_dict, session):
    timestamp = datetime.now().strftime('%Y/%m/%d/%H:%M')
    customer_ID = customer_name[:2].upper()  # Get the first two letters of Customer name in capitals
    policy_name = policy_data[sheet_headers.index('PolicyName')]
    policy_id = policy_data[sheet_headers.index('PolicyID')]
    cloud_type = policy_data[sheet_headers.index('CloudType')]
    p_section = str(policy_data[sheet_headers.index('SectionOption')]).strip()  # Ensure it's a string
    requirement_details = requirements_dict.get(p_section)
    print(f"Requirements for Section '{requirement_details}'")

    if not requirement_details:
        print(f"No requirement found for Section '{p_section}'")
        reason = f"Failed to update  - Missing Standard Section - " + timestamp
        return {'PolicyName': policy_name, 'PolicyID': policy_id, 'Reason': reason}
    else:
        requirement_name, compliance_id = requirement_details
    # Fetch existing policy data
    policy_url = f"{baseurl}/policy/{policy_id}"
    existing_policy_data = await make_get_request(policy_url, token_manager, session)
    if not existing_policy_data:
        print(f">Failed to fetch existing policy data for {policy_name} - {policy_id}")
        reason = f"Failed to fetch existing policy data " + timestamp
        return {'PolicyName': policy_name, 'PolicyID': policy_id, 'Reason': reason}
    else:
        # Update compliance metadata and labels
        existing_compliance_metadata = existing_policy_data.get('complianceMetadata', [])
        updated_compliance_metadata = existing_compliance_metadata[:]
        new_compliance_metadata = {
            "standardId": standard_id,
            "standardName": stnd_name,
            "requirementName": requirement_name,
            "sectionId": p_section,
            "policyId": policy_id,
            "complianceId": compliance_id,
            "customAssigned": "true"
        }
        # Remove unnecessary fields
        new_compliance_metadata.pop('sectionViewOrder', None)
        new_compliance_metadata.pop('requirementViewOrder', None)
        new_compliance_metadata.pop('systemDefault', None)
        if included_value == "YES":
            if not any(meta.get('standardId') == standard_id for meta in updated_compliance_metadata):
                updated_compliance_metadata.append(new_compliance_metadata)
        elif included_value == "REMOVE":
            updated_compliance_metadata = [meta for meta in updated_compliance_metadata if meta.get('standardId') != standard_id]
        labels = policy_data[sheet_headers.index('Labels')]

        if isinstance(labels, str):
            labels_list = [label.strip() for label in labels.split(',') if label.strip()]
        else:
            labels_list = labels if isinstance(labels, list) else []
        standard_label = stnd_name.replace(' ', '_')
        if included_value == "YES" and standard_label not in labels_list:
            labels_list.append(standard_label)
        p_section_label = str(p_section).replace(' ', '_')
        if included_value == "YES" and f'{customer_ID}_{cloud_type.upper()}_{p_section_label}' not in labels_list:
            labels_list.append(f'{customer_ID}_{cloud_type.upper()}_{p_section_label}')

        elif included_value == "REMOVE":
            labels_list = [label for label in labels_list if label not in [f'{customer_ID}_{cloud_type.upper()}_{p_section_label}', f'{standard_label}']]
        # Prepare payload and update policy
        # Extract only the required fields from the existing policy data
        original_payload = {
            "name": existing_policy_data.get("name", ""),
            "policyType": existing_policy_data.get("policyType", ""),
            "severity": existing_policy_data.get("severity", ""),
            "cloudType": existing_policy_data.get("cloudType", ""),
            "enabled": existing_policy_data.get("enabled", False)  # Default to False if not provided
        }
        # Process compliance metadata
        complliance_stnd_names = ', '.join(
            metadata.get('standardName', '') for metadata in existing_compliance_metadata
        )

        # Add the updated fields for complianceMetadata and labels
        payload = {
            **original_payload,  # Include only the original required fields
            "complianceMetadata": updated_compliance_metadata,  # Updated compliance metadata
            "labels": labels_list  # Updated labels
        }

        try:
            policy_id, response_text = await update_policy(baseurl, token_manager, policy_id, payload, session)
            if policy_id:
                return {'Included': included_value, 'PolicyName': policy_name, 'PolicyID': policy_id, 'Labels': labels_list,
                        'Standard': complliance_stnd_names, 'Reason': f"Updated " + timestamp}
            else:
                return {'PolicyName': policy_name, 'PolicyID': policy_id, 'Reason': f"Failed to update: {response_text} " + timestamp}
        except Exception as e:
            print(f">>Error updating policy {policy_name}: {str(e)}")
            return {'PolicyName': policy_name, 'PolicyID': policy_id, 'Reason': f"Failed to update: {str(e)} " + timestamp}


async def update_policy(baseurl, token_manager, policy_id, payload,  session):
    token = await token_manager.get_token()
    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }
    # print(f"Payload: {json.dumps(payload, indent=2)}")

    policy_url = f"{baseurl}/policy/{policy_id}"
    try:
        async with session.put(policy_url, headers=headers, json=payload, ssl=SSL_VERIFY) as response:
            response_text = await response.text()
            if response.status == 200:
                return policy_id, response_text
            else:
                print(f">>Failed to update policy {policy_id}. Status: {response.status}, Response: {response_text}")
                if response.status == '404':
                    print(f"Policy: {json.dumps(payload, indent=2)}")

                print(f"Policy: {policy_id} - {policy_name} does not exist, please add policy with setting status to YES")
                return None, response_text
    except Exception as e:
        print(f">>Failed to update policy {policy_id} - {policy_name}: {str(e)}")
        return None, str(e)


async def main():
    timestamp = datetime.now().strftime('%Y%m%d%H%M')

    test_event = {
        # "Customer": "SocGen",
        "Customer": "Awesome Inc",
        # "Customer": "Widget",
        "bucket_name": "data",
        "filename": "compliance_all_policies",
        # "standard_id": "581cc66e-4559-4cc6-be12-470d5d34077b", # PCS Awesome Standard
        # "standard_id": "df8fd24e-b25a-470f-b23e-ecb6d6766c87" # PSO MTSBv2_GCP_Mandatory Standard

        # "secret_name": "compliance-policy"
    }
    baseurl, access_key_id, secret_key = read_api_config()
    async with aiohttp.ClientSession() as session:
        token_manager = TokenManager(baseurl, access_key_id, secret_key, session)
        customer_name = test_event.get('Customer', 'Customer_Test')
        customer_ID = customer_name[:2].upper()  # Get the first two letters of Customer name in capitals
        standard_id = test_event.get('standard_id', '')
        customer_label = customer_name.replace(' ', '_')
        # standard_id = test_event.get('standard_id', 'b7a7ef28-ff11-4052-b9b7-a76e95fe902e')
        bucket_name = test_event.get('bucket_name', 'pc-compliance-pl')  # this is folder name
        filename = test_event.get('filename', 'compliance_policies')

        if standard_id:
            stnd_url = f"{baseurl}/compliance/{standard_id}"
            stnd_info = await make_get_request(stnd_url, token_manager, session)
            if not stnd_info:
                return []
            stnd_name = stnd_info.get('name', 'Unknown Standard')
            cloud_types = stnd_info.get("cloudType", [])
        else:
            standards_url = f"{baseurl}/compliance"
            standards_data = await make_get_request(standards_url, token_manager, session)
            # Filter the data for entries where "systemDefault" is False
            filtered_standards_data = [standard for standard in standards_data if not standard.get("systemDefault", True)]
            # Prompt the user to select a compliance standard
            selected_standard = await prompt_for_standard_selection(filtered_standards_data)
            if not selected_standard:
                print("No compliance standard selected.")
                return
            # Extract the stnd_id and cloud_types from the selected standard
            standard_id = selected_standard.get("id")
            cloud_types = selected_standard.get("cloudType", [])
            # cloud_types = ['azure', 'aws'] # Override manually if required
            stnd_name = selected_standard.get('name')
        stnd_label = stnd_name.replace(' ', '_')
        print(f"Selected Standard {stnd_name} - ID: {standard_id}")
        print(f"Cloud Types for this standard: {cloud_types}")
        print(f"Processing {customer_name} policies for {stnd_name}")
        # Get compliance metadata for the given standard_id
        compliance_url = f"{baseurl}/policy/compliance"
        compliance_metadata = await make_get_request(compliance_url, token_manager, session)
        input_file_key = f"{bucket_name}/{filename}_{customer_label}_{stnd_label}_latest.xlsx"
        consolidated_file_key = f"{bucket_name}/{filename}_{customer_label}_{stnd_label}_data_totals_{timestamp}.csv"
        updated_file_key = f"{bucket_name}/{filename}_{customer_label}_{stnd_label}_latest_{timestamp}.xlsx"
        standard_file_key = f"{bucket_name}/{filename}_{customer_label}_{stnd_label}_Reference_latest.csv"
        standard_summary_data = await generate_compliance_summary(baseurl, token_manager, standard_id, session)
        # print(standard_summary_data)
        if standard_summary_data:
            write_csv_to_file(standard_file_key, standard_summary_data)
        # Backup the input file to updated_file_key
        try:
            shutil.copy(input_file_key, updated_file_key)
            print(f"Backup created: {updated_file_key}")
        except FileNotFoundError:
            print(Fore.RED + f"Error: The file '{input_file_key}' does not exist. \nPolicies for {stnd_name} have not been collected yet. "
                             f"\n Please run the Compliance-All-Policies-excel.py script and select your chosen standard to generate initial policies spreadsheet and try again."+ Fore.WHITE)
        except Exception as e:
            print(Fore.RED + f"An unexpected error occurred while creating the backup: {e}")

        # Load workbook and get the sheet
        workbook = openpyxl.load_workbook(input_file_key)
        sheet = workbook.active
        # Collect existing data validation rules
        existing_validations = sheet.data_validations.dataValidation
        included_validation = None
        p_section_validation = None

        for dv in existing_validations:
            # Assuming the data validation for 'Included' and 'SectionOption'
            # were the ones originally added as list validations.
            # We might need a more robust way to identify them if there are other validations.
            if dv.type == 'list' and any(f'"{option}' in dv.formula1 for option in ["YES", "NO", "REVIEW", "MAYBE", "REMOVE"]):
                included_validation = dv
            elif dv.type == 'list' and dv.formula1.startswith('"'): # Assuming SectionOption list starts with a quote
                p_section_validation = dv

        # Perform your data updates here (as in your original first script)
        import_data = []
        for row in sheet.iter_rows(values_only=True):
            import_data.append(list(row))
        sheet_headers = import_data[0]
        rows = import_data[1:]
        # Extract index positions from headers
        try:
            policy_id_index = sheet_headers.index("PolicyID")
            section_option_index = sheet_headers.index('SectionOption')
            included_index = sheet_headers.index('Included')
        except ValueError as e:
            print(f"Error: One or more required headers not found: {e}")
            return

        existing_policy_ids = {row[policy_id_index] for row in rows}
        # Check and fetch new policies for relevant subtypes, might need to add certain policies if new subtypes are introduced
        new_policy_subtypes_to_check = [
            'run', 'run_and_build', 'build', 'audit', 'network_event', 'network',
            'ueba', 'dns', 'event', 'identity', 'permissions', 'misconfig', 'misconfig_and_event'
        ]
        new_policies = []
        sheet.title = f"{stnd_name} Policies"
        # Write headers to the Excel file
        for col_num, header in enumerate(sheet_headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        # Fetch all alerts counts at once
        alerts_count_dict = await get_all_alerts_count(baseurl, token_manager, session)
        alert_rules_dict = await get_policies_alert_rules(baseurl, token_manager, session)
        # Updates to existing rows
        print(f"Updating existing policies spreadsheet data with policy details from Prisma Cloud ---")
        await update_policies(baseurl, token_manager, rows, sheet_headers, alerts_count_dict, alert_rules_dict, cloud_types, session)
        # update_labels(rows, sheet_headers, customer_name, stnd_name) # For YES and MAYBE policies (optional)
        # Check for NEW YES Custom policies rows to process
        filtered_custom_rows = [
            row for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True))
            if row[included_index] == 'YES' and (row[sheet_headers.index('PolicyMode')] != 'redlock_default' or not row[policy_id_index])
        ]
        filtered_updated_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            included_value = row[included_index]
            # Custom YES results processed later
            # if included_value in ["REMOVE"] or (included_value in ["YES"] and row[sheet_headers.index('PolicyMode')] != 'custom' and row[sheet_headers.index('PolicyID')]):
            if included_value.upper() in ["REMOVE", "YES"] and row[policy_id_index]:
                # Any new Custom policies are not included here
                # print(f"Row meeting criteria: {row}")  # Debugging log
                filtered_updated_rows.append({
                    "row_data": row,  # Keep the full row if needed for further processing
                    "p_section": row[section_option_index]  # Extract the p_section from SectionOption column
                })


        # YES and REMOVE Policies Processing
        print(f"Updating YES and REMOVE policies for {stnd_name} ---")
        updated_policies = []
        failed_policies = []
        updated_results = []
        p_sections = []

        # Process the filtered rows
        if filtered_updated_rows:
            # Collect p_sections for the requirements lookup
            for item in filtered_updated_rows:
                p_section = item["p_section"]
                if p_section:  # Only include non-empty and non-None values
                    # print(f"Extracted p_section: {p_section}")
                    p_sections.append(p_section)
            # print(f"Collected p_sections: {p_sections}")  # Final debug log
            # Get requirements mapping for the collected p_sections
            requirements_dict = await get_requirements(baseurl, token_manager, standard_id, p_sections, p_section, session)
            # print(requirements_dict)
            print(f"Total YES/REMOVE Policies to update: {len(filtered_updated_rows)}")
            for item in filtered_updated_rows:
                row_data = item["row_data"]
                policy_name = row_data[sheet_headers.index('PolicyName')]
                included_value = row_data[included_index]
                if included_value in 'YES':
                    print(f">Adding to standard policy: {policy_name}")
                elif included_value in 'REMOVE':
                    print(f">Removing from standard policy: {policy_name}")
                # Process the updated policy
                result = await process_updated_policy(
                    baseurl, token_manager, row_data, sheet_headers, customer_name, included_value,
                    standard_id, stnd_name, requirements_dict, session
                )
                updated_results.append(result)
                if result and 'Updated' in result.get('Reason', ''):
                    updated_policies.append(result)
                elif result and 'Failed' in result.get('Reason', ''):
                    failed_policies.append(result)
            # Calculate update and failure counts
            updated_count = len(updated_policies)
            failed_count = len(failed_policies)
            # Print results
            print(f"YES/REMOVE Policies Updated: {updated_count}")
            # print(f"YES/REMOVE Updated Policies: {json.dumps(updated_policies, indent=2)}")
            print(f"YES/REMOVE Policies Failed: {failed_count}")
            print(f"YES/REMOVE Failed Policies: {json.dumps(failed_policies, indent=2)}")
            # Update the Excel sheet with results
            for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 to skip headers
                policy_name = sheet.cell(row=row_idx, column=sheet_headers.index('PolicyName') + 1).value
                result = next((r for r in updated_results if r and r.get('PolicyName') == policy_name), None)
                if result:
                    reason = result.get('Reason', 'No Reason Provided')
                    standard = result.get('Standard', '')
                    if reason not in 'Failed':
                        included_value_result = result.get('Included', '')
                        if included_value_result == "YES":
                            labels = ', '.join(result.get('Labels', [])) if isinstance(result.get('Labels', []), list) else result.get('Labels', '')
                            sheet.cell(row=row_idx, column=sheet_headers.index('Standard') + 1, value=stnd_name)
                            sheet.cell(row=row_idx, column=included_index + 1, value=customer_name)
                        else:
                            labels = ', '.join(
                                label for label in result.get('Labels', []) if isinstance(result.get('Labels', []), list) and not label.startswith(f'{customer_ID}_')
                            )
                            sheet.cell(row=row_idx, column=included_index + 1, value='NO')
                            sheet.cell(row=row_idx, column=sheet_headers.index('Standard') + 1, value=standard)
                            sheet.cell(row=row_idx, column=section_option_index + 1, value='')
                        sheet.cell(row=row_idx, column=sheet_headers.index('Labels') + 1, value=labels)
                    sheet.cell(row=row_idx, column=sheet_headers.index('Reason') + 1, value=reason)
        else:
            print(f"No YES/REMOVE Policies to update")

        # Filter custom policy rows for processing
        if filtered_custom_rows:
            print(f"Adding to standard Custom policies for {stnd_name} ---")
            # Get existing searches and policies
            search_history_url = f"{baseurl}/search/history?filter=saved"
            search_data = await make_get_request(search_history_url, token_manager, session)
            # Convert the list of searches into a dictionary for easy lookup
            search_data_dict = {search.get('searchName', ''): search.get('id', '') for search in search_data}
            custom_results = []
            for row in filtered_custom_rows:
                print(f">(Creating) and adding to standard Custom policy: {row[sheet_headers.index('PolicyName')]}")
                result = await process_custom_policy(
                    baseurl, token_manager, row, sheet_headers, customer_name, search_data_dict,
                    standard_id, stnd_name, compliance_metadata, session
                )
                custom_results.append(result)
            # Calculate counts for processed, created, and failed policies
            processed_count = len(custom_results)
            updated_count = sum(1 for r in custom_results if r and r.get('PolicyID') and 'updated' in r.get('Reason', ''))
            created_count = sum(1 for r in custom_results if r and r.get('PolicyID') and 'created' in r.get('Reason', ''))
            failed_count = sum(1 for r in custom_results if r and 'Failed' in r.get('Reason', ''))
            print(f"Custom Policies Processed: {processed_count}")
            print(f"Custom Policies Updated: {updated_count}")
            print(f"Custom Policies Created: {created_count}")
            print(f"Custom Policies Failed: {failed_count}")
            # Update the Excel with the custom_results
            # Start updating the Excel sheet with custom_results
            for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 to skip headers
                policy_name = sheet.cell(row=row_idx, column=sheet_headers.index('PolicyName') + 1).value
                result = next((r for r in custom_results if r and r.get('PolicyName') == policy_name), None)
                if result:
                    # Get values with default fallbacks
                    policy_id = result.get('PolicyID', '')
                    labels = ', '.join(result.get('Labels', [])) if isinstance(result.get('Labels', []), list) else result.get('Labels', '')
                    reason = result.get('Reason', 'No Reason Provided')
                    # Update the Excel sheet with PolicyID, Standard, Included, Labels, and Policy UPI
                    sheet.cell(row=row_idx, column=policy_id_index + 1, value=policy_id)
                    sheet.cell(row=row_idx, column=sheet_headers.index('Standard') + 1, value=stnd_name)
                    sheet.cell(row=row_idx, column=included_index + 1, value=customer_name)
                    sheet.cell(row=row_idx, column=sheet_headers.index('Labels') + 1, value=labels)
                    sheet.cell(row=row_idx, column=sheet_headers.index('Policy UPI') + 1, value='Custom')
                    # Update the Reason column
                    sheet.cell(row=row_idx, column=sheet_headers.index('Reason') + 1, value=reason)
                    # Update search_id and savedSearch if the policy is newly created
                    if reason not in 'Failed':
                        # Get the new values or retain the existing ones if not provided
                        search_id = result.get('search_id', sheet.cell(row=row_idx, column=sheet_headers.index('search_id') + 1).value)
                        saved_search = result.get('savedSearch', sheet.cell(row=row_idx, column=sheet_headers.index('savedSearch') + 1).value)
                        sheet.cell(row=row_idx, column=sheet_headers.index('search_id') + 1, value=search_id)
                        sheet.cell(row=row_idx, column=sheet_headers.index('savedSearch') + 1, value=saved_search)
                        # Update the Enabled column to 'TRUE' for new policies
                        sheet.cell(row=row_idx, column=sheet_headers.index('Enabled') + 1, value='TRUE')
        else:
            print(f"No Custom Policies to update")
        # Check for new policies and add them
        # Fetch and combine policies for each subtype
        for policy_subtype in new_policy_subtypes_to_check:
            policy_url = f"{baseurl}/v2/policy?policy.subtype={policy_subtype}"
            policies = await make_get_request(policy_url, token_manager, session)
            new_policies.extend(policies)  # Use extend to flatten the list
        # Filter out policies for the cloud types included in the standard that already exist
        new_policies = [
            policy for policy in new_policies
            # if policy["policyId"] not in existing_policy_ids and policy.get('cloudType') != 'alibaba_cloud'
            if policy["policyId"] not in existing_policy_ids and policy['cloudType'] in cloud_types
        ]
        if new_policies:
            print(f"Adding Newly discovered policies to spreadsheet data for {stnd_name}")
            # Process new policies and add them to rows
            print(f"Total new policies found: {len(new_policies)}")
            for policy in new_policies:
                await add_new_policy_to_rows(policy, baseurl, token_manager, rows, sheet_headers, alerts_count_dict, alert_rules_dict, session)
            error_if_included_yes.add(f"{section_option_col}2:{section_option_col}{sheet.max_row}")

        sheet.calculate_dimension()
        # Save the workbook
        workbook.save(input_file_key)
        print(f"All data written to {input_file_key} successfully.")
        # Process total numbers and save to csv
        totals, maybe_totals, yes_totals, custom_totals, alert_rule_totals, auto_dismiss_totals = process_consolidated_data(rows, sheet_headers, customer_name, stnd_name, standard_summary_data)
        csv_output = generate_csv(customer_name, stnd_name, totals, maybe_totals, yes_totals, custom_totals, alert_rule_totals, auto_dismiss_totals, alert_rules_dict)
        write_csv_to_file(consolidated_file_key, csv.reader(StringIO(csv_output)))
    # Record the end time
    end_time = time.time()
    # Calculate the elapsed time
    elapsed_time_seconds = end_time - start_time
    # Convert elapsed time to minutes
    elapsed_time_minutes = elapsed_time_seconds / 60
    # Print the script elapsed run time in minutes
    print(f"Time taken: {elapsed_time_minutes:.2f} minutes")
    return {
        'statusCode': 200,
        'body': json.dumps(f"CSV files created successfully as {input_file_key} and {consolidated_file_key}.")
    }


if __name__ == "__main__":
    asyncio.run(main())